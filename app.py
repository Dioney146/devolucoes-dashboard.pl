# ── COLUMN FIXES APPLIED based on real spreadsheet:
# ABA DEVOLUÇÕES (8456- DEVOLUCAO 2026):
#   VLTOTAL, DTENT (filter), DTENTREGA (delivery), NOTA_VENDA, NOTA_DEVOLUCAO,
#   NUMCAR, PLACA, DESTINO, MOTIVO, CODCLI, CLIENTE, MOTORISTA, NOMERCA,
#   NOMEFUNC, SUPERVISOR, TIPO_MERCADO, DTSAIDA, PRACA, NOME_CIDADE
# ABA REENTREGAS (8261 - REENTREGAS 2026):
#   VLTOTGER, DTRANSF, NUMTRANSVENDA, CODUSUR, TOTPESO, PLACAANT, PLACAATUAL,
#   MOTIVOTRANSF, CODMOTIVO, CLIENTE, NUMNOTA, NUMPED, PRACA, NOME (vendedor)
# ABA NOMES:
#   PLACA, MOTORISTA, ENTREGADOR  (cadastro fixo de equipe por veículo)
#
# ── REDESIGN VISUAL (v2) ─────────────────────────────────────────────────────
# Nenhuma regra de negócio, cálculo, filtro, consulta ou fonte de dados foi
# alterada. As mudanças são exclusivamente de apresentação e organização:
#   • Menu lateral fixo (navegação por páginas em vez de abas)
#   • Novo cabeçalho com status de sincronização, notificações e usuário
#   • Painel de filtros compacto + limpar filtros
#   • Cards de indicadores redesenhados
#   • Gráficos com eixos discretos, grid quase imperceptível e tooltip moderno
#   • Layout responsivo (desktop / notebook / tablet)
#
# ── v2.1 ─────────────────────────────────────────────────────────────────────
# Novo painel "Pedidos com devolução" no Dashboard: lista as notas de venda
# devolvidas no dia filtrado, com busca, ordenação e exportação em CSV.
# Usa a mesma base já filtrada (df) — nenhuma regra existente foi alterada.
#
# ── v2.2 (NOMES / equipe por placa) ──────────────────────────────────────────
# Nova fonte auxiliar: aba "NOMES" da mesma planilha (PLACA · MOTORISTA ·
# ENTREGADOR). A placa da base de devoluções é normalizada (sem espaços,
# hífens e em maiúsculas) e cruzada com esse cadastro, criando duas colunas
# derivadas: MOTORISTA_NOMES e ENTREGADOR_NOMES.
# O que foi acrescentado:
#   • Gráficos "Devoluções por motorista" e "Devoluções por entregador"
#     (valor em barras + quantidade de notas na linha)
#   • Nome do motorista/entregador no tooltip do gráfico por placa
#   • Colunas MOTORISTA e ENTREGADOR na lista de pedidos e na página Campos
#   • Filtro global por motorista
#   • Nova página "Equipe" com ranking completo por motorista/entregador
# Nenhum cálculo, filtro ou consulta já existente foi alterado — o cruzamento
# é apenas um enriquecimento de colunas.

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import requests
import io
from datetime import datetime, date

st.set_page_config(
    page_title="Gestão de Devoluções Delly's",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ═════════════════════════════════════════════════════════════════════════════
# DESIGN SYSTEM
# ═════════════════════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Sora:wght@300;400;500;600;700&family=Inter:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500;700&display=swap');

:root{
  --bg-0:#04070f; --bg-1:#070c18; --bg-2:#0b1424;
  --panel:rgba(14,23,41,0.62); --panel-solid:rgba(10,17,32,0.94);
  --line:rgba(120,170,225,0.13); --line-strong:rgba(56,189,248,0.30);
  --cyan:#22d3ee; --blue:#3b82f6; --sky:#38bdf8;
  --violet:#a78bfa; --green:#34d399; --amber:#fbbf24; --red:#f87171;
  --txt-0:#f1f6fc; --txt-1:#b9c8dc; --txt-2:#7c8ea8; --txt-3:#4e5f78;
  --r-lg:20px; --r-md:14px; --r-sm:10px;
}

*,*::before,*::after{box-sizing:border-box;}
html,body,.stApp{font-family:'Inter',sans-serif;color:var(--txt-1);background:var(--bg-0);}

/* ── Fundo espacial discreto ─────────────────────────────────────────────── */
.bg-overlay{position:fixed;inset:0;z-index:0;pointer-events:none;}
.bg-img{position:absolute;inset:0;
  background-image:url('https://images.unsplash.com/photo-1450101499163-c8848c66ca85?w=1920&q=80');
  background-size:cover;background-position:center;
  filter:blur(7px) brightness(0.16) saturate(0.35);transform:scale(1.1);opacity:0.55;}
.bg-tint{position:absolute;inset:0;
  background:
    radial-gradient(1100px 620px at 18% -8%,rgba(34,211,238,0.09),transparent 62%),
    radial-gradient(900px 560px at 88% 4%,rgba(167,139,250,0.07),transparent 60%),
    linear-gradient(180deg,rgba(4,7,15,0.90) 0%,rgba(4,7,15,0.955) 55%,rgba(4,7,15,0.98) 100%);}

.stApp,[data-testid="stAppViewContainer"],[data-testid="stHeader"]{background:transparent!important;}
[data-testid="stAppViewContainer"]>section{background:transparent!important;}
#MainMenu,footer,header{visibility:hidden!important;display:none!important;}
.stDeployButton,[data-testid="stStatusWidget"],[data-testid="stToolbar"],
[data-testid="stHeader"],[data-testid="stDecoration"]{display:none!important;}
.main .block-container{position:relative;z-index:1;padding-top:1.1rem!important;
  padding-bottom:3rem!important;padding-left:1.6rem!important;padding-right:1.6rem!important;max-width:100%!important;}

/* ── Menu lateral ────────────────────────────────────────────────────────── */
section[data-testid="stSidebar"]{
  background:linear-gradient(180deg,rgba(7,12,24,0.99),rgba(9,16,31,0.99))!important;
  border-right:1px solid var(--line)!important;box-shadow:1px 0 40px rgba(0,0,0,0.5);}
section[data-testid="stSidebar"] .block-container{padding-top:1.4rem!important;}
section[data-testid="stSidebar"] [data-testid="stSidebarCollapseButton"]{opacity:.35;}
.side-brand{display:flex;align-items:center;gap:12px;padding:2px 6px 18px;
  border-bottom:1px solid var(--line);margin-bottom:16px;}
.side-mark{width:40px;height:40px;border-radius:12px;flex:0 0 40px;
  background:linear-gradient(140deg,#0891b2,#2563eb);display:flex;align-items:center;
  justify-content:center;font-size:19px;
  box-shadow:0 0 0 1px rgba(56,189,248,0.28),0 6px 22px rgba(8,145,178,0.35);}
.side-brand-t{font-family:'Sora',sans-serif;font-size:0.86rem;font-weight:600;
  color:var(--txt-0);line-height:1.15;letter-spacing:.01em;}
.side-brand-s{font-size:0.6rem;color:var(--txt-3);letter-spacing:.16em;
  text-transform:uppercase;font-weight:600;margin-top:3px;}
.side-cap{font-size:0.6rem;color:var(--txt-3);letter-spacing:.2em;text-transform:uppercase;
  font-weight:700;margin:14px 0 8px 8px;}

section[data-testid="stSidebar"] div[role="radiogroup"]{gap:2px!important;}
section[data-testid="stSidebar"] div[role="radiogroup"] label{
  width:100%;padding:9px 12px!important;border-radius:var(--r-sm)!important;
  border:1px solid transparent!important;transition:all .18s ease;cursor:pointer;
  background:transparent!important;margin:0!important;}
section[data-testid="stSidebar"] div[role="radiogroup"] label:hover{
  background:rgba(56,189,248,0.06)!important;}
section[data-testid="stSidebar"] div[role="radiogroup"] label>div:first-child{display:none!important;}
section[data-testid="stSidebar"] div[role="radiogroup"] label p{
  font-size:0.83rem!important;font-weight:500!important;color:var(--txt-2)!important;
  letter-spacing:.005em;margin:0!important;}
section[data-testid="stSidebar"] div[role="radiogroup"] label:has(input:checked){
  background:linear-gradient(90deg,rgba(34,211,238,0.16),rgba(37,99,235,0.07))!important;
  border-color:rgba(56,189,248,0.30)!important;
  box-shadow:inset 3px 0 0 var(--cyan),0 4px 18px rgba(8,145,178,0.14);}
section[data-testid="stSidebar"] div[role="radiogroup"] label:has(input:checked) p{
  color:#e6fbff!important;font-weight:600!important;}
.side-foot{margin-top:20px;padding:12px 12px;border-top:1px solid var(--line);
  font-size:0.66rem;color:var(--txt-3);line-height:1.7;}

/* ── Cabeçalho ───────────────────────────────────────────────────────────── */
.topbar{display:flex;align-items:center;justify-content:space-between;gap:22px;
  background:linear-gradient(100deg,rgba(9,15,29,0.90),rgba(11,20,38,0.86));
  border:1px solid var(--line);border-radius:var(--r-lg);
  padding:16px 24px;margin-bottom:16px;
  backdrop-filter:blur(20px);box-shadow:0 10px 44px rgba(0,0,0,0.42);}
.tb-left{display:flex;align-items:center;gap:16px;min-width:0;}
.tb-icon{width:46px;height:46px;flex:0 0 46px;border-radius:13px;
  background:linear-gradient(140deg,#0e7490,#1d4ed8);display:flex;align-items:center;
  justify-content:center;font-size:21px;
  box-shadow:0 0 0 1px rgba(56,189,248,0.25),0 8px 26px rgba(14,165,233,0.28);}
.tb-title{font-family:'Sora',sans-serif;font-size:1.16rem;font-weight:600;
  color:var(--txt-0);letter-spacing:.015em;line-height:1.2;margin:0;}
.tb-sub{font-size:0.63rem;color:var(--txt-3);font-weight:600;letter-spacing:.19em;
  text-transform:uppercase;margin:5px 0 0;}
.tb-right{display:flex;align-items:center;gap:10px;flex-wrap:wrap;justify-content:flex-end;}
.tb-chip{display:flex;align-items:center;gap:8px;padding:8px 13px;border-radius:999px;
  background:rgba(255,255,255,0.035);border:1px solid var(--line);
  font-size:0.72rem;color:var(--txt-1);font-weight:500;white-space:nowrap;}
.tb-chip .dot{width:7px;height:7px;border-radius:50%;background:var(--green);
  box-shadow:0 0 9px rgba(52,211,153,0.85);}
.tb-chip.bell{position:relative;padding:8px 12px;font-size:0.85rem;}
.tb-badge{position:absolute;top:2px;right:4px;min-width:16px;height:16px;padding:0 4px;
  border-radius:999px;background:var(--red);color:#2a0606;font-size:0.6rem;font-weight:700;
  display:flex;align-items:center;justify-content:center;border:2px solid var(--bg-1);}
.tb-user{display:flex;align-items:center;gap:10px;padding:6px 14px 6px 6px;border-radius:999px;
  background:rgba(255,255,255,0.035);border:1px solid var(--line);}
.tb-av{width:31px;height:31px;border-radius:50%;background:linear-gradient(140deg,#22d3ee,#6366f1);
  display:flex;align-items:center;justify-content:center;font-size:0.72rem;font-weight:700;color:#04121a;}
.tb-un{font-size:0.75rem;color:var(--txt-0);font-weight:600;line-height:1.1;}
.tb-ur{font-size:0.62rem;color:var(--txt-3);letter-spacing:.08em;text-transform:uppercase;margin-top:2px;}

/* ── Painéis / cards ─────────────────────────────────────────────────────── */
.panel{background:linear-gradient(158deg,rgba(15,24,42,0.72),rgba(8,13,25,0.66));
  border:1px solid var(--line);border-radius:var(--r-lg);
  padding:22px 24px 16px;backdrop-filter:blur(16px);
  box-shadow:0 14px 46px rgba(0,0,0,0.38),inset 0 1px 0 rgba(255,255,255,0.03);
  margin-bottom:18px;}
.panel .js-plotly-plot{margin-top:2px;}
.panel-h{display:flex;align-items:center;justify-content:space-between;gap:14px;
  margin-bottom:14px;}
.panel-t{display:flex;align-items:center;gap:11px;}
.panel-t .bar{width:4px;height:24px;border-radius:2px;
  background:linear-gradient(180deg,var(--cyan),var(--blue));box-shadow:0 0 11px rgba(34,211,238,0.5);}
.panel-t h3{font-family:'Sora',sans-serif;font-size:1.14rem;font-weight:600;color:var(--txt-0);
  margin:0;letter-spacing:.01em;}
.panel-tag{font-size:0.76rem;color:var(--txt-3);letter-spacing:.13em;text-transform:uppercase;
  font-weight:600;padding:5px 11px;border-radius:999px;border:1px solid var(--line);
  background:rgba(255,255,255,0.025);white-space:nowrap;}

/* ── KPIs ────────────────────────────────────────────────────────────────── */
.kpi-grid{display:grid;grid-template-columns:repeat(5,minmax(0,1fr));gap:14px;margin-bottom:20px;}
.kpi{position:relative;overflow:hidden;border-radius:var(--r-lg);padding:18px 19px 16px;
  background:linear-gradient(150deg,rgba(16,26,46,0.82),rgba(10,17,32,0.72));
  border:1px solid var(--line);backdrop-filter:blur(14px);
  transition:transform .22s ease,border-color .22s ease,box-shadow .22s ease;}
.kpi:hover{transform:translateY(-3px);border-color:var(--line-strong);
  box-shadow:0 14px 40px rgba(8,145,178,0.14);}
.kpi::after{content:'';position:absolute;left:18px;right:18px;bottom:0;height:1px;
  background:linear-gradient(90deg,transparent,var(--acc,var(--cyan)),transparent);opacity:.55;}
.kpi-top{display:flex;align-items:center;justify-content:space-between;margin-bottom:13px;}
.kpi-ico{width:31px;height:31px;border-radius:9px;display:flex;align-items:center;
  justify-content:center;font-size:0.92rem;background:rgba(255,255,255,0.045);
  border:1px solid var(--line);}
.kpi-delta{font-family:'JetBrains Mono',monospace;font-size:0.66rem;font-weight:700;
  padding:3px 8px;border-radius:999px;letter-spacing:.02em;}
.kpi-up{color:#fecaca;background:rgba(248,113,113,0.14);border:1px solid rgba(248,113,113,0.28);}
.kpi-down{color:#bbf7d0;background:rgba(52,211,153,0.13);border:1px solid rgba(52,211,153,0.28);}
.kpi-flat{color:var(--txt-2);background:rgba(255,255,255,0.04);border:1px solid var(--line);}
.kpi-val{font-family:'JetBrains Mono',monospace;font-size:1.92rem;font-weight:700;
  color:var(--acc,var(--cyan));line-height:1.05;letter-spacing:-0.02em;
  white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
.kpi-lab{font-size:0.66rem;color:var(--txt-1);font-weight:700;letter-spacing:.14em;
  text-transform:uppercase;margin:9px 0 4px;}
.kpi-sub{font-size:0.66rem;color:var(--txt-3);}

/* ── Filtros ─────────────────────────────────────────────────────────────── */
.filters{background:linear-gradient(120deg,rgba(12,20,38,0.80),rgba(9,15,29,0.74));
  border:1px solid var(--line);border-radius:var(--r-lg);padding:14px 20px 4px;
  margin-bottom:18px;backdrop-filter:blur(16px);}
.filters-h{display:flex;align-items:center;gap:9px;font-size:0.63rem;color:var(--txt-2);
  letter-spacing:.17em;text-transform:uppercase;font-weight:700;margin-bottom:6px;}
.filters-h .pip{width:6px;height:6px;border-radius:50%;background:var(--cyan);
  box-shadow:0 0 9px rgba(34,211,238,0.8);}

label,.stSelectbox label,.stMultiSelect label,.stTextInput label,.stRadio label p{
  color:var(--txt-2)!important;font-size:0.68rem!important;font-weight:600!important;
  letter-spacing:.11em!important;text-transform:uppercase!important;}
.stTextInput input,.stDateInput input{background:rgba(255,255,255,0.035)!important;
  border:1px solid var(--line)!important;border-radius:var(--r-sm)!important;
  color:var(--txt-0)!important;font-family:'Inter',sans-serif!important;font-size:0.85rem!important;}
.stTextInput input:focus{border-color:var(--line-strong)!important;
  box-shadow:0 0 0 3px rgba(34,211,238,0.10)!important;}
.stSelectbox div[data-baseweb="select"]>div,.stMultiSelect div[data-baseweb="select"]>div{
  background:rgba(255,255,255,0.035)!important;border:1px solid var(--line)!important;
  border-radius:var(--r-sm)!important;color:var(--txt-0)!important;}
.stMultiSelect span[data-baseweb="tag"]{background:rgba(34,211,238,0.15)!important;
  color:#a5f3fc!important;border-radius:6px!important;}
div[data-baseweb="popover"] ul{background:var(--panel-solid)!important;
  border:1px solid var(--line)!important;}

.stButton>button{background:rgba(255,255,255,0.04)!important;color:var(--txt-1)!important;
  border:1px solid var(--line)!important;border-radius:var(--r-sm)!important;
  font-weight:600!important;font-size:0.8rem!important;padding:9px 18px!important;
  transition:all .2s ease!important;}
.stButton>button:hover{border-color:var(--line-strong)!important;color:var(--txt-0)!important;
  background:rgba(56,189,248,0.09)!important;}
.stButton>button[kind="primary"]{
  background:linear-gradient(120deg,#0891b2,#2563eb)!important;color:#ecfeff!important;
  border:1px solid rgba(56,189,248,0.45)!important;
  box-shadow:0 6px 22px rgba(8,145,178,0.34)!important;}
.stButton>button[kind="primary"]:hover{transform:translateY(-1px);
  box-shadow:0 10px 30px rgba(8,145,178,0.46)!important;}
.stDownloadButton>button{background:rgba(52,211,153,0.09)!important;color:#6ee7b7!important;
  border:1px solid rgba(52,211,153,0.28)!important;border-radius:var(--r-sm)!important;
  font-weight:600!important;}

/* ── Card do gráfico comparativo ─────────────────────────────────────────── */
div[data-testid="stVerticalBlockBorderWrapper"]:has(.cc-head){
  background:linear-gradient(155deg,rgba(13,22,40,0.72),rgba(8,14,27,0.66))!important;
  border:1px solid rgba(56,189,248,0.16)!important;border-radius:22px!important;
  padding:22px 26px 18px!important;margin-bottom:18px;
  backdrop-filter:blur(18px);
  box-shadow:0 18px 60px rgba(0,0,0,0.45),inset 0 1px 0 rgba(255,255,255,0.035);}
.cc-head{display:flex;align-items:flex-start;justify-content:space-between;gap:26px;
  flex-wrap:wrap;padding-bottom:16px;margin-bottom:6px;
  border-bottom:1px solid rgba(120,170,225,0.09);}
.cc-head-l{min-width:240px;}
.cc-title{font-family:'Sora',sans-serif;font-size:1.02rem;font-weight:600;color:#f1f6fc;
  letter-spacing:.01em;margin:0;line-height:1.25;}
.cc-sub{font-size:0.76rem;color:#7c8ea8;margin:7px 0 0;font-weight:400;}
.cc-head-r{display:flex;align-items:stretch;gap:12px;flex-wrap:wrap;}
.cc-mini{display:flex;flex-direction:column;justify-content:center;gap:3px;
  min-width:132px;padding:11px 16px;border-radius:14px;
  background:rgba(255,255,255,0.032);border:1px solid rgba(120,170,225,0.13);}
.cc-mini-lab{font-size:0.6rem;color:#6d8099;font-weight:700;letter-spacing:.15em;
  text-transform:uppercase;}
.cc-mini-val{font-family:'JetBrains Mono',monospace;font-size:1.16rem;font-weight:700;
  letter-spacing:-0.02em;line-height:1.1;}
.cc-mini-sub{font-size:0.63rem;color:#4e5f78;}
.cc-ctrls{display:flex;align-items:center;gap:8px;}
.cc-pick{font-size:0.73rem;color:#b9c8dc;padding:9px 14px;border-radius:11px;
  background:rgba(255,255,255,0.03);border:1px solid rgba(120,170,225,0.13);white-space:nowrap;}
.cc-dots{font-size:1.05rem;color:#6d8099;padding:6px 11px;border-radius:11px;
  background:rgba(255,255,255,0.03);border:1px solid rgba(120,170,225,0.13);line-height:1.1;}
.cc-foot{display:flex;align-items:center;gap:8px;flex-wrap:wrap;margin-top:4px;
  padding:15px 20px;border-radius:16px;
  background:linear-gradient(120deg,rgba(255,255,255,0.035),rgba(255,255,255,0.015));
  border:1px solid rgba(120,170,225,0.11);}
.cc-foot-i{display:flex;flex-direction:column;gap:5px;flex:1 1 170px;min-width:150px;}
.cc-foot-lab{display:flex;align-items:center;gap:7px;font-size:0.66rem;color:#6d8099;
  font-weight:700;letter-spacing:.13em;text-transform:uppercase;}
.cc-foot-lab .d{width:7px;height:7px;border-radius:50%;display:inline-block;}
.cc-foot-val{font-family:'JetBrains Mono',monospace;font-size:1.24rem;font-weight:700;
  letter-spacing:-0.02em;}
.cc-foot-sep{width:1px;align-self:stretch;
  background:linear-gradient(180deg,transparent,rgba(120,170,225,0.18),transparent);}
@media (max-width:900px){
  .cc-head{gap:16px;}
  .cc-head-r{width:100%;}
  .cc-mini{flex:1 1 140px;min-width:0;}
  .cc-ctrls{display:none;}
  .cc-foot-sep{display:none;}
  .cc-foot-val{font-size:1.06rem;}
  div[data-testid="stVerticalBlockBorderWrapper"]:has(.cc-head){padding:18px 16px 14px!important;}
}

/* ── Tabelas ─────────────────────────────────────────────────────────────── */
.tbl-wrap{background:rgba(6,11,22,0.86);border:1px solid var(--line);
  border-radius:var(--r-md);overflow:auto;max-height:540px;}
.tbl-wrap table{width:100%;border-collapse:collapse;}
.tbl-wrap th{position:sticky;top:0;background:rgba(11,20,38,0.99);color:var(--sky);
  font-size:0.68rem;font-weight:700;letter-spacing:.11em;text-transform:uppercase;
  text-align:left;padding:11px 14px;white-space:nowrap;border-bottom:1px solid var(--line-strong);}
.tbl-wrap td{padding:9px 14px;color:var(--txt-1);font-size:0.8rem;white-space:nowrap;
  border-bottom:1px solid rgba(120,170,225,0.06);}
.tbl-wrap tr:hover td{background:rgba(56,189,248,0.05);}
.num{font-family:'JetBrains Mono',monospace;font-weight:600;}

/* ── Diversos ────────────────────────────────────────────────────────────── */
hr{border-color:var(--line)!important;margin:20px 0!important;}
.stAlert{background:rgba(12,22,42,0.8)!important;border:1px solid var(--line)!important;
  border-radius:var(--r-md)!important;}
.stCaption,[data-testid="stCaptionContainer"]{color:var(--txt-3)!important;font-size:.72rem!important;}
.stTabs [data-baseweb="tab-list"]{background:rgba(9,16,31,0.8)!important;border-radius:var(--r-md)!important;
  padding:4px!important;gap:4px!important;border:1px solid var(--line)!important;}
.stTabs [data-baseweb="tab"]{background:transparent!important;border-radius:var(--r-sm)!important;
  color:var(--txt-2)!important;font-weight:600!important;font-size:0.8rem!important;}
.stTabs [aria-selected="true"]{background:rgba(34,211,238,0.13)!important;color:#e6fbff!important;}
.streamlit-expanderHeader,details summary{color:var(--txt-2)!important;font-size:0.8rem!important;}
::-webkit-scrollbar{width:8px;height:8px;}
::-webkit-scrollbar-track{background:transparent;}
::-webkit-scrollbar-thumb{background:rgba(56,189,248,0.20);border-radius:8px;}
::-webkit-scrollbar-thumb:hover{background:rgba(56,189,248,0.38);}

/* ── Tipografia dos gráficos (ajuste responsivo) ─────────────────────────── */
@media (max-width:1400px){
  .js-plotly-plot .xtick text,.js-plotly-plot .ytick text{font-size:12px!important;}
  .js-plotly-plot .legend text{font-size:13px!important;}
  .js-plotly-plot .textpoint text,.js-plotly-plot .bartext{font-size:14px!important;}
}
@media (max-width:1000px){
  .js-plotly-plot .xtick text,.js-plotly-plot .ytick text{font-size:11px!important;}
  .js-plotly-plot .legend text{font-size:12px!important;}
  .js-plotly-plot .textpoint text,.js-plotly-plot .bartext{font-size:12px!important;}
  .panel-t h3{font-size:1rem;}
}

/* ── Responsivo ──────────────────────────────────────────────────────────── */
@media (max-width:1500px){ .kpi-val{font-size:1.62rem;} }
@media (max-width:1200px){ .kpi-val{font-size:1.44rem;} }
@media (max-width:1200px){
  .kpi-grid{grid-template-columns:repeat(3,minmax(0,1fr));}
  .main .block-container{padding-left:1.2rem!important;padding-right:1.2rem!important;}
  .tb-title{font-size:1.02rem;}
}
@media (max-width:820px){
  .kpi-grid{grid-template-columns:repeat(2,minmax(0,1fr));}
  .topbar{flex-direction:column;align-items:flex-start;gap:14px;}
  .tb-right{justify-content:flex-start;}
}
@media (prefers-reduced-motion:reduce){
  .kpi,.stButton>button{transition:none!important;}
  .kpi:hover{transform:none;}
}
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="bg-overlay"><div class="bg-img"></div><div class="bg-tint"></div></div>',
            unsafe_allow_html=True)


# ═════════════════════════════════════════════════════════════════════════════
# HELPERS DE FORMATAÇÃO E GRÁFICOS  (regras de cálculo inalteradas)
# ═════════════════════════════════════════════════════════════════════════════
def fmt_brl(v):
    try:
        return f"R$ {float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return "R$ 0,00"


def fmt_brl0(v):
    """Formata valor monetário em R$ SEM casas decimais (arredondado), usado nos gráficos."""
    try:
        return f"R$ {round(float(v)):,}".replace(",", ".")
    except:
        return "R$ 0"


HOVER = dict(bgcolor="rgba(6,10,18,0.97)", bordercolor="rgba(56,189,248,0.28)",
             font=dict(color="#e8f2ff", family="Inter", size=13), align="left")

# ═════════════════════════════════════════════════════════════════════════════
# EXPORTAÇÃO — Excel já formatado como tabela (e CSV como alternativa)
# ═════════════════════════════════════════════════════════════════════════════
# O conteúdo exportado é exatamente o que está na tela, com os mesmos filtros.
# A conversão abaixo só devolve aos números o formato numérico (o app mostra
# "R$ 1.234,56" como texto), para que a planilha permita somar e ordenar.
_RE_MOEDA = r"^R\$\s?-?[\d.]+,\d{2}$"
_RE_PCT = r"^-?[\d.]+,?\d*%$"


def _coluna_numerica(serie, padrao):
    vals = serie.astype(str).str.strip()
    nao_vazias = vals[vals != ""]
    if len(nao_vazias) == 0:
        return False
    return bool(nao_vazias.str.match(padrao).all())


def to_xlsx(df_exp, sheet_name="Dados"):
    """Gera um .xlsx com tabela nomeada, filtro, cabeçalho fixo e larguras.

    Devolve None se o openpyxl não estiver disponível no ambiente, para o app
    seguir funcionando apenas com o CSV.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except Exception:
        return None

    df_x = df_exp.copy()
    df_x.columns = [str(c) for c in df_x.columns]
    formatos = {}

    for col in df_x.columns:
        # is_numeric_dtype cobre pandas 2 (object) e pandas 3 (dtype str)
        if not pd.api.types.is_numeric_dtype(df_x[col]):
            if _coluna_numerica(df_x[col], _RE_MOEDA):
                df_x[col] = (df_x[col].astype(str).str.replace("R$", "", regex=False)
                             .str.strip().str.replace(".", "", regex=False)
                             .str.replace(",", ".", regex=False))
                df_x[col] = pd.to_numeric(df_x[col], errors="coerce")
                formatos[col] = 'R$ #,##0.00'
            elif _coluna_numerica(df_x[col], _RE_PCT):
                df_x[col] = (df_x[col].astype(str).str.replace("%", "", regex=False)
                             .str.replace(".", "", regex=False)
                             .str.replace(",", ".", regex=False))
                df_x[col] = pd.to_numeric(df_x[col], errors="coerce") / 100
                formatos[col] = '0.0%'

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "Dados"

    cab_fill = PatternFill("solid", fgColor="0F2135")
    cab_font = Font(color="E8F2FF", bold=True, size=11, name="Calibri")
    borda = Border(bottom=Side(style="thin", color="2C4763"))

    ws.append(list(df_x.columns))
    for c in range(1, len(df_x.columns) + 1):
        cel = ws.cell(row=1, column=c)
        cel.fill, cel.font, cel.border = cab_fill, cab_font, borda
        cel.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    for _, linha in df_x.iterrows():
        ws.append([None if pd.isna(v) else v for v in linha.tolist()])

    for i, col in enumerate(df_x.columns, start=1):
        letra = get_column_letter(i)
        _larg_dados = df_x[col].astype(str).str.len().max() if len(df_x) else 0
        if pd.isna(_larg_dados):
            _larg_dados = 0
        largura = max(len(str(col)) + 4, int(_larg_dados) + 3)
        ws.column_dimensions[letra].width = min(max(largura, 10), 48)
        if col in formatos:
            for r in range(2, len(df_x) + 2):
                ws.cell(row=r, column=i).number_format = formatos[col]

    if len(df_x) > 0:
        ref = f"A1:{get_column_letter(len(df_x.columns))}{len(df_x) + 1}"
        tab = Table(displayName="Tabela1", ref=ref)
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(tab)
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_tabela(df_exp, nome_base, key, sheet_name="Dados", largura_total=False):
    """Renderiza os botões de exportação: Excel (tabela pronta) e CSV."""
    if df_exp is None or len(df_exp) == 0:
        return
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    xlsx = to_xlsx(df_exp, sheet_name)
    csv = df_exp.to_csv(index=False, sep=";", decimal=",").encode("utf-8-sig")

    if xlsx is None:
        st.download_button("Exportar CSV", data=csv, file_name=f"{nome_base}_{stamp}.csv",
                           mime="text/csv", key=f"{key}_csv",
                           use_container_width=largura_total)
        st.caption("Para exportar em Excel, adicione `openpyxl` ao requirements.txt.")
        return

    b1, b2 = st.columns(2, gap="small")
    with b1:
        st.download_button(
            "Exportar tabela (Excel)", data=xlsx,
            file_name=f"{nome_base}_{stamp}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"{key}_xlsx", use_container_width=True)
    with b2:
        st.download_button("Exportar CSV", data=csv, file_name=f"{nome_base}_{stamp}.csv",
                           mime="text/csv", key=f"{key}_csv", use_container_width=True)


# ═════════════════════════════════════════════════════════════════════════════
# DESIGN TOKENS DOS GRÁFICOS — identidade única para todo o sistema
# ═════════════════════════════════════════════════════════════════════════════
# Tipografia: Inter para rótulos e categorias, JetBrains Mono para números.
F_TXT = "Inter"
F_NUM = "JetBrains Mono"

# Linguagem de cor por SIGNIFICADO (nunca por gráfico):
#   crítico/alto volume → vermelho · intermediário → laranja
#   quantidade/indicador → amarelo · baixo volume/informativo → azul
#   positivo/concluído → verde · neutro → cinza
C_CRIT = "#ef4444"     # vermelho — atenção, alto volume
C_WARN = "#f97316"     # laranja  — intermediário
C_QTY = "#facc15"      # amarelo  — quantidade / indicador (linhas)
C_INFO = "#38bdf8"     # azul     — baixo volume, informativo
C_OK = "#22c55e"       # verde    — positivo, concluído
C_NEUT = "#8fa3bb"     # cinza    — neutro, comparação histórica
C_BASE = C_INFO

# Pontos e realces derivados (versões claras, usadas em marcadores)
P_QTY, P_INFO, P_OK, P_NEUT = "#fde68a", "#bae6fd", "#bbf7d0", "#cbd5e1"

# Superfícies e traços
TXT_HI = "#f8fafc"     # números principais
TXT_MID = "#c3d0e0"    # categorias
TXT_LOW = "#7d90a8"    # apoio
GRID = "rgba(148,180,220,0.05)"
AXIS = "rgba(148,180,220,0.15)"
BADGE = "rgba(4,8,15,0.78)"

# Escalas contínuas mantidas por compatibilidade com as chamadas existentes.
BLUE = ["#0b3a5c", "#0d7fb8", "#38bdf8", "#a5dcf8"]
RED = ["#5c1414", "#b91c1c", "#ef4444", "#fca5a5"]
GREEN = ["#0f3d24", "#15803d", "#22c55e", "#a7f3c4"]
MIXED = [C_CRIT, C_WARN, C_QTY, C_INFO, C_OK, C_NEUT, "#22d3ee", "#fb923c"]
VIOLET = ["#312150", "#6d3fc0", "#a78bfa", "#ddd2fb"]


def ramp(n, top=5, mid=10, base=C_BASE):
    """Cor por criticidade: os primeiros colocados sempre em vermelho/laranja."""
    return [C_CRIT if i < top else C_WARN if i < mid else base for i in range(n)]


def trunc(v, n=26):
    """Encurta rótulos longos preservando a leitura (o nome cheio fica no hover)."""
    s = str(v)
    return s if len(s) <= n else s[: n - 1] + "…"


def eixo_x(labels, size=13, angle=-38):
    """Eixo X técnico: rótulos encurtados, sem grade, linha de base discreta."""
    return dict(
        tickmode="array", tickvals=list(labels), ticktext=[trunc(v, 22) for v in labels],
        tickfont=dict(color=TXT_MID, size=size, family=F_NUM),
        gridcolor="rgba(0,0,0,0)", linecolor=AXIS, zeroline=False,
        tickangle=angle, automargin=True, ticks="outside", ticklen=4, tickcolor=AXIS)


def eixo_y_oculto(topo=None):
    """Eixo Y sem escala lateral — os valores ficam nos próprios elementos."""
    d = dict(showticklabels=False, gridcolor=GRID, zeroline=False, linecolor="rgba(0,0,0,0)")
    if topo:
        d["range"] = [0, topo]
    return d


def base_layout(height, margin, legenda=False):
    """Camada comum a todos os gráficos: fundo transparente sobre o grafite do painel."""
    lay = dict(
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        font=dict(color=TXT_MID, family=F_TXT, size=13),
        height=height, margin=margin, separators=",.",
        hoverlabel=HOVER, hovermode="closest", coloraxis_showscale=False,
        showlegend=legenda,
    )
    if legenda:
        lay["legend"] = dict(bgcolor="rgba(0,0,0,0)", bordercolor="rgba(0,0,0,0)",
                             font=dict(color=TXT_MID, size=13, family=F_TXT),
                             orientation="h", x=0, xanchor="left", y=1.10,
                             itemsizing="constant", itemwidth=40, tracegroupgap=20)
    return lay


def plotly_dark(fig, height=None, margin_b=40):
    """Compatibilidade: aplica a base do sistema a figuras criadas fora dos helpers."""
    fig.update_layout(**base_layout(height or 360, dict(t=16, b=margin_b, l=8, r=12)))
    fig.update_xaxes(showticklabels=False, gridcolor=GRID, zeroline=False, linecolor=AXIS)
    fig.update_yaxes(showticklabels=False, gridcolor=GRID, zeroline=False, linecolor=AXIS)
    return fig


def panel_open(title, tag=None, icon=""):
    tg = f'<span class="panel-tag">{tag}</span>' if tag else ""
    st.markdown(
        f'<div class="panel"><div class="panel-h"><div class="panel-t"><div class="bar"></div>'
        f'<h3>{icon} {title}</h3></div>{tg}</div>', unsafe_allow_html=True)


def panel_close():
    st.markdown('</div>', unsafe_allow_html=True)


def html_table(df_in, accent="#38bdf8", max_rows=500, min_width=900):
    heads = "".join([f"<th>{c}</th>" for c in df_in.columns])
    rws = ""
    for _, row in df_in.head(max_rows).iterrows():
        cells = "".join([f"<td>{v}</td>" for v in row.values])
        rws += f"<tr>{cells}</tr>"
    st.markdown(
        f'<div class="tbl-wrap"><table style="min-width:{min_width}px;">'
        f'<thead><tr>{heads}</tr></thead><tbody>{rws}</tbody></table></div>',
        unsafe_allow_html=True)


# ═════════════════════════════════════════════════════════════════════════════
# FONTE DE DADOS  (devoluções e reentregas inalteradas · NOMES é nova)
# ═════════════════════════════════════════════════════════════════════════════
SHEET_ID = "1GCw6vE5lrIZYJUKnQlKvBMX71CgIdxcRBA1YCrjFadI"
GSHEETS_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv&id={SHEET_ID}"

REENTREGAS_URLS = [
    f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv&sheet=8261+-+REENTREGAS+2026",
    f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv&sheet=8261%20-%20REENTREGAS%202026",
    f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv&sheet=REENTREGAS+2026",
    f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv&sheet=REENTREGAS",
    f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv&gid=1",
    f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv&gid=2",
]

# ── Aba NOMES (cadastro de equipe por placa) ────────────────────────────────
# O endpoint gviz costuma ser o mais confiável para buscar uma aba pelo nome.
# As demais variações ficam como alternativa caso a planilha esteja publicada
# em outro formato.
NOMES_URLS = [
    f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/gviz/tq?tqx=out:csv&sheet=NOMES",
    f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/gviz/tq?tqx=out:csv&sheet=Nomes",
    f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv&sheet=NOMES",
    f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv&sheet=Nomes",
]


@st.cache_data(ttl=60)
def load_data(url):
    r = requests.get(url, timeout=30, headers={"User-Agent": "Mozilla/5.0"})
    r.raise_for_status()
    return pd.read_csv(io.StringIO(r.text))


@st.cache_data(ttl=60)
def load_reentregas():
    """Tenta múltiplas URLs até encontrar a aba de reentregas com dados válidos."""
    erros = []
    for url in REENTREGAS_URLS:
        try:
            r = requests.get(url, timeout=30, headers={"User-Agent": "Mozilla/5.0"})
            r.raise_for_status()
            df = pd.read_csv(io.StringIO(r.text))
            cols_up = [str(c).strip().upper().replace(" ", "_") for c in df.columns]
            if any(c in cols_up for c in ["DTRANSF", "MOTIVOTRANSF", "NUMNOTA", "VLTOTGER"]):
                return df, url, None
            else:
                erros.append(f"URL ok mas colunas não reconhecidas: {cols_up[:6]}")
        except Exception as e:
            erros.append(f"{url.split('sheet=')[-1][:40]} → {str(e)[:80]}")
    return None, None, erros


@st.cache_data(ttl=60)
def load_nomes():
    """Carrega a aba NOMES (PLACA · MOTORISTA · ENTREGADOR).

    Só aceita a resposta quando encontra uma coluna de placa — assim evita
    trazer por engano a primeira aba da planilha.
    """
    erros = []
    for url in NOMES_URLS:
        try:
            r = requests.get(url, timeout=30, headers={"User-Agent": "Mozilla/5.0"})
            r.raise_for_status()
            df = pd.read_csv(io.StringIO(r.text))
            cols_up = [str(c).strip().upper().replace(" ", "_") for c in df.columns]
            if "PLACA" in cols_up and any(c in cols_up for c in ["MOTORISTA", "ENTREGADOR"]):
                return df, url, None
            erros.append(f"URL respondeu mas sem PLACA/MOTORISTA: {cols_up[:6]}")
        except Exception as e:
            erros.append(f"{url.split('sheet=')[-1][:30]} → {str(e)[:80]}")
    return None, None, erros


def parse_brl(s):
    s = str(s).replace("R$", "").strip()
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    return pd.to_numeric(s, errors="coerce")


def norm_placa(v):
    """Normaliza a placa para o cruzamento: sem espaços, hífens ou pontos."""
    return (str(v).upper().strip()
            .replace("-", "").replace(".", "").replace(" ", "").replace("/", ""))


with st.spinner("Carregando dados..."):
    try:
        df_raw = load_data(GSHEETS_URL)
    except Exception as e:
        st.error(f"Não foi possível carregar as devoluções: {e}")
        st.stop()

df_reent_raw = None
reent_load_error = None
reent_url_usada = None
try:
    df_reent_raw, reent_url_usada, _erros = load_reentregas()
    if df_reent_raw is None:
        reent_load_error = "Nenhuma URL funcionou: " + " | ".join(_erros or [])
except Exception as e:
    reent_load_error = str(e)

df_nomes = None
nomes_load_error = None
nomes_url_usada = None
try:
    df_nomes, nomes_url_usada, _erros_nomes = load_nomes()
    if df_nomes is None:
        nomes_load_error = "Nenhuma URL funcionou: " + " | ".join(_erros_nomes or [])
except Exception as e:
    nomes_load_error = str(e)

# ── Normaliza colunas devoluções ────────────────────────────────────────────
df_raw.columns = [str(c).strip().upper().replace(" ", "_") for c in df_raw.columns]
actual_cols = list(df_raw.columns)


def get_col(df, names):
    for n in names:
        if n in df.columns:
            return n
    return None


VALOR_COL = get_col(df_raw, ["VLTOTAL", "VLT", "VL_TOTAL", "VALOR_LIQUIDO", "VALOR", "TOTAL"]) or "VLTOTAL"
COL_PLACA = get_col(df_raw, ["PLACA"])
COL_MOTIVO = get_col(df_raw, ["MOTIVO", "MOTIVO_DEVOLUCAO", "MOTIVO_DEV"])
COL_CLIENTE = get_col(df_raw, ["CLIENTE", "NOME_CLIENTE", "RAZAO_SOCIAL"])
COL_VENDEDOR = get_col(df_raw, ["NOMERCA", "VENDEDOR", "NOME_VENDEDOR"])
COL_DEVOLUCION = get_col(df_raw, ["NOMEFUNC", "DEVOLUCIONISTA", "FUNCIONARIO"])
COL_MOTORISTA = get_col(df_raw, ["MOTORISTA", "ENTREGADOR"])
COL_DESTINO = get_col(df_raw, ["DESTINO", "NOME_CIDADE", "CIDADE", "MUNICIPIO"])
COL_NF_VENDA = get_col(df_raw, ["NOTA_VENDA", "NF_VENDA", "NF_SAIDA", "NOTA_SAIDA", "NOTA_FISCAL"])
COL_NOTA_DEV = get_col(df_raw, ["NOTA_DEVOLUCAO", "NF_DEVOLUCAO"])
COL_NUMCAR = get_col(df_raw, ["NUMCAR", "NUM_CARREGAMENTO", "CARREGAMENTO"])
COL_CODCLI = get_col(df_raw, ["CODCLI", "COD_CLI", "CLI"])
COL_SUPERVISOR = get_col(df_raw, ["SUPERVISOR", "AM", "GERENTE"])
COL_PRACA = get_col(df_raw, ["PRACA"])
COL_TIPO_MERC = get_col(df_raw, ["TIPO_MERCADO", "CANAL", "SEGMENTO"])
COL_DTSAIDA = get_col(df_raw, ["DTSAIDA", "DATA_DEVOLUCAO", "DATA", "DT_DEVOLUCAO"])
COL_DTENTREGA = get_col(df_raw, ["DTENT", "DTENTREGA", "DATA_ENTREGA", "DT_ENTREGA"])
# Número do pedido, quando a planilha o disponibilizar (opcional)
COL_NUMPED = get_col(df_raw, ["NUMPED", "NUM_PED", "PEDIDO", "NUMPEDIDO"])

if VALOR_COL not in df_raw.columns:
    df_raw[VALOR_COL] = 0.0

df_raw[VALOR_COL] = df_raw[VALOR_COL].apply(parse_brl).fillna(0)
for col in df_raw.columns:
    if col != VALOR_COL:
        df_raw[col] = df_raw[col].fillna("").astype(str).str.strip()

if COL_DTENTREGA:
    df_raw["_DTENTREGA_DT"] = pd.to_datetime(df_raw[COL_DTENTREGA], dayfirst=True, errors="coerce")
    mask_nat = df_raw["_DTENTREGA_DT"].isna() & (df_raw[COL_DTENTREGA] != "")
    if mask_nat.any():
        df_raw.loc[mask_nat, "_DTENTREGA_DT"] = pd.to_datetime(
            df_raw.loc[mask_nat, COL_DTENTREGA], format="%Y-%m-%d", errors="coerce")
else:
    df_raw["_DTENTREGA_DT"] = pd.NaT


# ═════════════════════════════════════════════════════════════════════════════
# CRUZAMENTO COM A ABA NOMES — placa → motorista / entregador
# ═════════════════════════════════════════════════════════════════════════════
# Enriquecimento puro: apenas acrescenta duas colunas em df_raw. Nenhuma linha
# é removida e nenhum cálculo existente muda. Placas sem cadastro ficam com
# "NÃO CADASTRADO" para continuarem visíveis nos gráficos.
COL_MOT_NOME = "MOTORISTA_NOMES"
COL_ENT_NOME = "ENTREGADOR_NOMES"
SEM_CADASTRO = "NÃO CADASTRADO"

mapa_motorista = {}
mapa_entregador = {}
df_nomes_norm = None

if df_nomes is not None and len(df_nomes) > 0:
    df_nomes = df_nomes.copy()
    df_nomes.columns = [str(c).strip().upper().replace(" ", "_") for c in df_nomes.columns]

    _c_placa_n = get_col(df_nomes, ["PLACA", "PLACAS", "VEICULO"])
    _c_mot_n = get_col(df_nomes, ["MOTORISTA", "NOME_MOTORISTA", "MOT"])
    _c_ent_n = get_col(df_nomes, ["ENTREGADOR", "AJUDANTE", "NOME_ENTREGADOR", "AJUD"])

    if _c_placa_n:
        for _c in df_nomes.columns:
            df_nomes[_c] = df_nomes[_c].fillna("").astype(str).str.strip()
        df_nomes["_PLACA_KEY"] = df_nomes[_c_placa_n].apply(norm_placa)
        df_nomes = df_nomes[df_nomes["_PLACA_KEY"] != ""]
        # Se a mesma placa aparecer mais de uma vez, vale o último cadastro.
        if _c_mot_n:
            mapa_motorista = dict(zip(df_nomes["_PLACA_KEY"], df_nomes[_c_mot_n]))
        if _c_ent_n:
            mapa_entregador = dict(zip(df_nomes["_PLACA_KEY"], df_nomes[_c_ent_n]))
        df_nomes_norm = df_nomes
    else:
        nomes_load_error = "A aba NOMES foi lida, mas não tem uma coluna PLACA."

if COL_PLACA:
    df_raw["_PLACA_KEY"] = df_raw[COL_PLACA].apply(norm_placa)
else:
    df_raw["_PLACA_KEY"] = ""

df_raw[COL_MOT_NOME] = (df_raw["_PLACA_KEY"].map(mapa_motorista)
                        .fillna("").replace("", SEM_CADASTRO).str.upper().str.strip())
df_raw[COL_ENT_NOME] = (df_raw["_PLACA_KEY"].map(mapa_entregador)
                        .fillna("").replace("", SEM_CADASTRO).str.upper().str.strip())
df_raw.loc[df_raw[COL_MOT_NOME] == "", COL_MOT_NOME] = SEM_CADASTRO
df_raw.loc[df_raw[COL_ENT_NOME] == "", COL_ENT_NOME] = SEM_CADASTRO

# Percentual de placas efetivamente cruzadas (usado no diagnóstico)
placas_base = set(df_raw.loc[df_raw["_PLACA_KEY"] != "", "_PLACA_KEY"])
placas_cadastradas = set(mapa_motorista.keys()) | set(mapa_entregador.keys())
placas_sem_cadastro = sorted(placas_base - placas_cadastradas)

# ── Normaliza colunas reentregas ────────────────────────────────────────────
REENT_COLS_MAP = {
    "NUMNOTA": ["NUMNOTA"],
    "DTFAT": ["DTFAT"],
    "SERIE": ["SERIE"],
    "ESPECIE": ["ESPECIE"],
    "DTSAIDA": ["DTSAIDA"],
    "VLTOTGER": ["VLTOTGER", "VLTOT", "VLTOTAL"],
    "TOTPESO": ["TOTPESO", "PESO"],
    "NUMTRANSVENDA": ["NUMTRANSVENDA", "NUMTRANS"],
    "NUMCARANTERIOR": ["NUMCARANTERIOR"],
    "PLACAANT": ["ANTERIOR", "PLACAANT", "PLACA_ANT", "PLACAANTERIOR"],
    "COD_MOT_ANTERIOR": ["COD_MOT_ANTERIOR"],
    "NOME_MOT_ANTERIOR": ["NOME_MOT_ANTERIOR"],
    "COD_AJU_ANTERIOR": ["COD_AJU_ANTERIOR"],
    "NOME_AJU_ANTERIOR": ["NOME_AJU_ANTERIOR"],
    "NUMCARATUAL": ["NUMCARATUAL"],
    "PLACAATUAL": ["PLACA_ATUAL", "PLACAATUAL", "PLACA_ATU"],
    "COD_MOT_ATUAL": ["COD_MOT_ATUAL"],
    "NOME_MOT_ATUAL": ["NOME_MOT_ATUAL"],
    "COD_AJU_ATUAL": ["COD_AJU_ATUAL"],
    "NOME_AJU_ATUAL": ["NOME_AJU_ATUAL"],
    "DATATRANSF": ["DTRANSF", "DATATRANSF", "DATA_TRANSF"],
    "CODMOTIVO": ["CODMOTIVO"],
    "MOTIVOTRANSF": ["MOTIVOTRANSF", "MOTIVO_TRANSF"],
    "CODCLI": ["CODCLI"],
    "CLIENTE": ["CLIENTE"],
    "BAIRROENT": ["BAIRROENT"],
    "CODPRACA": ["CODPRACA"],
    "PRACA": ["PRACA"],
    "ROTA": ["ROTA"],
    "NUMPED": ["NUMPED"],
    "CODUSU": ["CODUSUR", "CODUSU"],
    "NOME": ["NOME", "RNOME"],
}

df_reent = None
reent_cols = {}
REENT_VALOR_COL = "_VALOR_REENT"

if df_reent_raw is not None:
    df_reent_raw.columns = [str(c).strip().upper().replace(" ", "_") for c in df_reent_raw.columns]

    def get_col_reent(alts):
        for n in alts:
            if n.strip().upper().replace(" ", "_") in df_reent_raw.columns:
                return n.strip().upper().replace(" ", "_")
        return None

    for canonical, alts in REENT_COLS_MAP.items():
        reent_cols[canonical] = get_col_reent(alts)

    rv = get_col_reent(["VLTOTGER", "VLTOT", "VLTOTAL", "VALOR", "TOTAL"])
    if rv:
        REENT_VALOR_COL = rv
        df_reent_raw[rv] = df_reent_raw[rv].apply(parse_brl).fillna(0)
    else:
        df_reent_raw["_VALOR_REENT"] = 0.0

    for col in df_reent_raw.columns:
        if col != REENT_VALOR_COL:
            df_reent_raw[col] = df_reent_raw[col].fillna("").astype(str).str.strip()

    dt_col_r = reent_cols.get("DATATRANSF")
    if not dt_col_r:
        for _c in ["DTRANSF", "DATATRANSF", "DATA_TRANSF"]:
            if _c in df_reent_raw.columns:
                dt_col_r = _c
                reent_cols["DATATRANSF"] = _c
                break
    if dt_col_r and dt_col_r in df_reent_raw.columns:
        df_reent_raw["_DATATRANSF_DT"] = pd.to_datetime(df_reent_raw[dt_col_r], dayfirst=True, errors="coerce")
        m2 = df_reent_raw["_DATATRANSF_DT"].isna() & (df_reent_raw[dt_col_r] != "")
        if m2.any():
            df_reent_raw.loc[m2, "_DATATRANSF_DT"] = pd.to_datetime(
                df_reent_raw.loc[m2, dt_col_r], format="%Y-%m-%d", errors="coerce")
    else:
        df_reent_raw["_DATATRANSF_DT"] = pd.NaT

    df_reent = df_reent_raw.copy()


# ═════════════════════════════════════════════════════════════════════════════
# CABEÇALHO
# ═════════════════════════════════════════════════════════════════════════════
_sync = datetime.now().strftime("%d/%m/%Y às %H:%M")
_n_reent = len(df_reent) if df_reent is not None else 0
_n_nomes = len(df_nomes_norm) if df_nomes_norm is not None else 0

st.markdown(f"""
<div class="topbar">
  <div class="tb-left">
    <div class="tb-icon">📦</div>
    <div>
      <p class="tb-title">GESTÃO DE DEVOLUÇÕES DELLY'S</p>
      <p class="tb-sub">Módulo de análise e controle operacional</p>
    </div>
  </div>
  <div class="tb-right">
    <div class="tb-chip"><span class="dot"></span> Sincronizado {_sync}</div>
    <div class="tb-chip bell">🔔<span class="tb-badge">{min(_n_reent, 99)}</span></div>
    <div class="tb-user">
      <div class="tb-av">DL</div>
      <div>
        <div class="tb-un">Delly's Logística</div>
        <div class="tb-ur">Roteirização</div>
      </div>
    </div>
    <div class="tb-chip">⚙️ Perfil</div>
  </div>
</div>
""", unsafe_allow_html=True)


# ═════════════════════════════════════════════════════════════════════════════
# MENU LATERAL
# ═════════════════════════════════════════════════════════════════════════════
PAGINAS = [
    "📊  Dashboard",
    "🧑‍✈️  Equipe",
    "🔄  Reentregas",
    "🔍  Detalhes Reentregas",
    "🗂️  Campos",
    "📑  Dados Completos",
    "👥  Clientes",
    "❗  Motivos",
    "🚚  Veículos",
    "📤  Relatórios",
    "⚙️  Configurações",
]

with st.sidebar:
    st.markdown("""
    <div class="side-brand">
      <div class="side-mark">📦</div>
      <div>
        <div class="side-brand-t">Delly's Food Service</div>
        <div class="side-brand-s">Logística</div>
      </div>
    </div>
    <div class="side-cap">Navegação</div>
    """, unsafe_allow_html=True)

    pagina = st.radio("Navegação", PAGINAS, label_visibility="collapsed", key="nav")

    st.markdown(f"""
    <div class="side-foot">
      Devoluções carregadas: <b style="color:#38bdf8;">{len(df_raw)}</b><br>
      Reentregas carregadas: <b style="color:#34d399;">{_n_reent}</b><br>
      Placas cadastradas (NOMES): <b style="color:#a78bfa;">{_n_nomes}</b><br>
      Cache: 60s
    </div>
    """, unsafe_allow_html=True)

pagina = pagina.split("  ", 1)[-1].strip()


# ═════════════════════════════════════════════════════════════════════════════
# FILTROS GLOBAIS (devoluções) — mesma lógica de filtragem do sistema atual
# ═════════════════════════════════════════════════════════════════════════════
PAGS_COM_FILTRO = {"Dashboard", "Equipe", "Campos", "Dados Completos", "Clientes",
                   "Motivos", "Veículos", "Relatórios"}

usar_data = False
dt_sel = None
sel_dev = []
sel_motivo = []
sel_mot_nome = []

if pagina in PAGS_COM_FILTRO:
    st.markdown('<div class="filters"><div class="filters-h"><span class="pip"></span>Filtros — devoluções</div>',
                unsafe_allow_html=True)
    fc1, fc2, fc3, fc6, fc4, fc5 = st.columns([2.6, 2.2, 2.2, 2.2, 1.3, 1], gap="medium")

    with fc1:
        datas_ok = df_raw["_DTENTREGA_DT"].dropna()
        if len(datas_ok) > 0:
            datas_unicas = sorted(datas_ok.dt.date.unique())
            opcoes_data = ["— Todas as datas —"] + [d.strftime("%d/%m/%Y") for d in datas_unicas]
            sel_data_str = st.selectbox("Data", opcoes_data, key="g_dtsel")
            if sel_data_str != "— Todas as datas —":
                dt_sel = datetime.strptime(sel_data_str, "%d/%m/%Y").date()
                usar_data = True
        else:
            st.caption("Sem datas válidas na coluna de data.")

    with fc2:
        if COL_DEVOLUCION:
            devs_opts = sorted([x for x in df_raw[COL_DEVOLUCION].unique() if x not in ("", "N/D", "nan", "None")])
            sel_dev = st.multiselect("Devolucionista", devs_opts, default=[], key="g_dev", placeholder="Todos")

    with fc3:
        if COL_MOTIVO:
            mot_opts = sorted([x for x in df_raw[COL_MOTIVO].unique() if x not in ("", "N/D", "nan", "None")])
            sel_motivo = st.multiselect("Motivo", mot_opts, default=[], key="g_mot", placeholder="Todos")

    with fc6:
        # Filtro novo, alimentado pela aba NOMES
        nomes_opts = sorted([x for x in df_raw[COL_MOT_NOME].unique() if x not in ("", "nan", "None")])
        sel_mot_nome = st.multiselect("Motorista", nomes_opts, default=[], key="g_motnome",
                                      placeholder="Todos")

    with fc4:
        st.markdown("<div style='height:26px'></div>", unsafe_allow_html=True)
        if st.button("Atualizar dados", use_container_width=True, type="primary", key="btn_upd_main"):
            st.cache_data.clear()
            st.rerun()

    with fc5:
        st.markdown("<div style='height:26px'></div>", unsafe_allow_html=True)
        if st.button("Limpar", use_container_width=True, key="btn_clear"):
            for k in ("g_dtsel", "g_dev", "g_mot", "g_motnome"):
                st.session_state.pop(k, None)
            st.rerun()

    st.markdown('</div>', unsafe_allow_html=True)

    if nomes_load_error:
        st.warning(f"Aba NOMES não carregada — motoristas e entregadores ficarão como "
                   f"'{SEM_CADASTRO}'. Detalhe: {nomes_load_error[:180]}")

# ── Aplica filtros (lógica idêntica à original + motorista) ─────────────────
df = df_raw.copy()
if usar_data and dt_sel:
    df = df[df["_DTENTREGA_DT"].dt.date == dt_sel]
if sel_dev and COL_DEVOLUCION:
    df = df[df[COL_DEVOLUCION].isin(sel_dev)]
if sel_motivo and COL_MOTIVO:
    df = df[df[COL_MOTIVO].isin(sel_motivo)]
if sel_mot_nome:
    df = df[df[COL_MOT_NOME].isin(sel_mot_nome)]

total_val = df[VALOR_COL].sum()
total_notas = len(df)
total_clientes = df[COL_CLIENTE].nunique() if COL_CLIENTE else 0
ticket_medio = total_val / total_notas if total_notas > 0 else 0
total_placas = df[COL_PLACA].nunique() if COL_PLACA else 0
total_motoristas = df[df[COL_MOT_NOME] != SEM_CADASTRO][COL_MOT_NOME].nunique()
total_entregadores = df[df[COL_ENT_NOME] != SEM_CADASTRO][COL_ENT_NOME].nunique()

if pagina in PAGS_COM_FILTRO:
    filtros_info = []
    if usar_data and dt_sel:
        filtros_info.append(f"Data {dt_sel.strftime('%d/%m/%Y')}")
    if sel_dev:
        filtros_info.append(f"Devolucionista: {', '.join(sel_dev[:2])}{'…' if len(sel_dev) > 2 else ''}")
    if sel_motivo:
        filtros_info.append(f"{len(sel_motivo)} motivo(s)")
    if sel_mot_nome:
        filtros_info.append(f"Motorista: {', '.join(sel_mot_nome[:2])}{'…' if len(sel_mot_nome) > 2 else ''}")
    if filtros_info:
        st.info(f"{' · '.join(filtros_info)} — **{total_notas} registros filtrados**")


# ═════════════════════════════════════════════════════════════════════════════
# FUNÇÕES DE GRÁFICO
# ═════════════════════════════════════════════════════════════════════════════
def anotar_linha(fig, xs, ys, ref_pxs, plot_h, frac_scale, cor, size,
                 yaxis="y2", ocupados=None, gap=30, fmt=lambda v: f"<b>{int(v)}</b>"):
    """Posiciona os números de uma linha evitando colisões, em pixels.

    ref_pxs  : altura (px) já ocupada por outros rótulos naquele x (ex.: o valor
               da barra). O rótulo da linha se afasta quando ficaria em cima.
    ocupados : lista acumulada de (x, px) de rótulos já posicionados, para que
               séries diferentes no mesmo gráfico também não se sobreponham.

    Cada número recebe uma etiqueta escura semitransparente, para nunca se
    confundir com a barra, a linha ou o fundo. Nada aqui altera os dados.
    """
    if ocupados is None:
        ocupados = []
    CANDIDATOS = [(26, 0), (-28, 0), (26, 30), (26, -30), (-28, 30), (-28, -30),
                  (42, 0), (-44, 0), (58, 0), (74, 0), (90, 0)]
    for i, (x, y) in enumerate(zip(xs, ys)):
        ponto_px = float(y) * frac_scale * plot_h
        conflitos = [px for xx, px, dx in ocupados if xx == x]
        if ref_pxs is not None:
            conflitos.extend(ref_pxs[i] if isinstance(ref_pxs[i], list) else [ref_pxs[i]])

        dy, dx = CANDIDATOS[0]
        for cy, cx in CANDIDATOS:
            alvo = ponto_px + cy
            if alvo < 14 or alvo > plot_h - 16:      # respeita as bordas do gráfico
                continue
            livre = all(abs(alvo - c) >= gap or cx != 0 for c in conflitos)
            livre = livre and all(abs(alvo - px) >= gap or cx != ddx
                                  for xx, px, ddx in ocupados if xx == x)
            if livre:
                dy, dx = cy, cx
                break

        ocupados.append((x, ponto_px + dy, dx))
        fig.add_annotation(
            x=x, y=y, xref="x", yref=yaxis, text=fmt(y), showarrow=False,
            yshift=dy, xshift=dx,
            font=dict(color=cor, size=size, family=F_NUM),
            bgcolor=BADGE, borderpad=4, bordercolor="rgba(148,180,220,0.16)",
            borderwidth=1, opacity=0.98,
        )
    return ocupados


def make_combo_chart(df_data, x_col, val_col, qtd_col, title, periodo="", bar_colors=None,
                     linha_nome="Quantidade (notas)", linha_cor=C_QTY,
                     linha_ponto=P_QTY, linha_texto=C_QTY, linha_rotulo="Notas",
                     customdata=None, extra_hover=""):
    """Combinado padrão do sistema: barras = valor (R$) · linha = quantidade.

    Barras sólidas com cor por criticidade, valor em branco negrito acima de
    cada barra e a linha em amarelo (a cor de quantidade em todo o sistema),
    com os números em etiquetas escuras que nunca cobrem os valores das barras.
    """
    n = len(df_data)
    if bar_colors is None:
        bar_colors = ramp(n)
    fig = go.Figure()

    max_val = df_data[val_col].max() if len(df_data) > 0 else 1
    max_qtd = df_data[qtd_col].max() if len(df_data) > 0 else 1

    _hb = "<b>%{x}</b>" + extra_hover + "<br>Valor: %{text}<extra></extra>"
    _hl = "<b>%{x}</b>" + extra_hover + "<br>" + linha_rotulo + ": %{y}<extra></extra>"

    # Tamanhos que encolhem conforme a quantidade de categorias, para o texto
    # nunca invadir a barra vizinha nem sair da área do gráfico.
    fs_val = 17 if n <= 14 else 15 if n <= 22 else 13
    fs_lin = 17 if n <= 14 else 15 if n <= 22 else 13
    fs_cat = 14 if n <= 14 else 12 if n <= 26 else 11

    fig.add_trace(go.Bar(
        x=df_data[x_col], y=df_data[val_col], name="Valor (R$)",
        marker=dict(color=bar_colors, opacity=1,
                    line=dict(color="rgba(255,255,255,0.10)", width=1)),
        width=0.52,
        text=[fmt_brl0(v) for v in df_data[val_col]],
        textposition="outside", cliponaxis=False,
        textfont=dict(size=fs_val, color=TXT_HI, family=F_NUM),
        customdata=customdata, hovertemplate=_hb, yaxis="y1",
    ))
    # Halo suave sob a linha, para separá-la das barras sem virar neon.
    fig.add_trace(go.Scatter(
        x=df_data[x_col], y=df_data[qtd_col], mode="lines", showlegend=False,
        line=dict(color=linha_cor, width=14, shape="spline"),
        opacity=0.10, hoverinfo="skip", yaxis="y2",
    ))
    fig.add_trace(go.Scatter(
        x=df_data[x_col], y=df_data[qtd_col], name=linha_nome,
        mode="lines+markers",
        line=dict(color=linha_cor, width=2.4, shape="spline"),
        marker=dict(color=linha_ponto, size=10, line=dict(color=linha_cor, width=2)),
        customdata=customdata, hovertemplate=_hl, yaxis="y2",
    ))

    h = max(620, min(n * 60, 1000))
    mt, mb = 70, 118
    fig.update_layout(**base_layout(h, dict(t=mt, b=mb, l=8, r=14), legenda=True))
    fig.update_layout(
        bargap=0.42, bargroupgap=0.08,
        title=dict(text=(f"<span style='font-size:13px;color:{TXT_LOW}'>{periodo}</span>"
                         if periodo else ""), x=0.5, xanchor="center", y=0.985),
        xaxis=eixo_x(df_data[x_col], size=fs_cat),
        yaxis=dict(**eixo_y_oculto(max_val * 1.42), side="left"),
        yaxis2=dict(showticklabels=False, overlaying="y", side="right", showgrid=False,
                    zeroline=False, range=[0, max_qtd * 2.9]),
    )

    # ── Rótulos da linha: posicionados dinamicamente para não colidir ───────
    plot_h = h - mt - mb
    ref_barra = [float(v) / (max_val * 1.42) * plot_h + 22 if max_val > 0 else 22
                 for v in df_data[val_col]]
    anotar_linha(fig, list(df_data[x_col]), list(df_data[qtd_col]),
                 ref_pxs=ref_barra, plot_h=plot_h,
                 frac_scale=(1 / (max_qtd * 2.9)) if max_qtd > 0 else 0,
                 cor=linha_texto, size=fs_lin, gap=32)
    return fig


def make_bar_simple(df_data, x_col, y_col, bar_colors, title_txt="", ylabel="Qtd"):
    """Barras verticais de quantidade — valor em branco negrito sobre a barra."""
    n = len(df_data)
    fs_val = 15 if n <= 14 else 13 if n <= 24 else 11
    fs_cat = 13 if n <= 14 else 11 if n <= 26 else 10
    _max = float(df_data[y_col].max()) if n else 1

    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=df_data[x_col], y=df_data[y_col],
        marker=dict(color=bar_colors[:n], opacity=1,
                    line=dict(color="rgba(255,255,255,0.08)", width=1)),
        width=0.55,
        text=[f"<b>{v}</b>" for v in df_data[y_col]], textposition="outside",
        cliponaxis=False,
        textfont=dict(size=fs_val, color=TXT_HI, family=F_NUM),
        hovertemplate="<b>%{x}</b><br>" + ylabel + ": %{y}<extra></extra>",
    ))
    h = max(400, min(n * 42, 620))
    fig.update_layout(**base_layout(h, dict(t=34 if title_txt else 22, b=96, l=10, r=16)))
    fig.update_layout(
        bargap=0.42,
        title=dict(text=(f"<span style='font-size:12px;color:{TXT_LOW}'>{title_txt}</span>"
                         if title_txt else ""), x=0.5, xanchor="center"),
        xaxis=eixo_x(df_data[x_col], size=fs_cat),
        yaxis=eixo_y_oculto(_max * 1.22),
    )
    return fig


def make_hbar(df_data, x_col, y_col, color_scale=None, height=400, money=True):
    """Ranking horizontal — leitura imediata do primeiro ao último colocado.

    Os dados chegam em ordem crescente (o topo do gráfico é o 1º colocado).
    A cor segue a mesma linguagem do sistema: os maiores em vermelho/laranja,
    os demais em azul. `color_scale` é aceito por compatibilidade e ignorado,
    para que nenhum gráfico invente uma paleta própria.
    """
    n = len(df_data)
    # posição no ranking: 0 = maior valor (última linha do dataframe)
    pos = list(range(n))[::-1]
    cores = [C_CRIT if p == 0 else C_WARN if p <= 2 else C_INFO for p in pos]

    fs_val = 15 if n <= 10 else 13
    fs_cat = 13 if n <= 10 else 12
    textos = [fmt_brl0(v) if money else f"<b>{int(v)}</b>" for v in df_data[x_col]]
    # margem à direita proporcional ao maior rótulo, para o número nunca cortar
    mr = max(76, min(int(max(len(t) for t in textos) * 9.5) + 26, 190))

    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=df_data[x_col], y=[trunc(v, 30) for v in df_data[y_col]], orientation="h",
        marker=dict(color=cores, opacity=1,
                    line=dict(color="rgba(255,255,255,0.08)", width=1)),
        text=textos, textposition="outside", cliponaxis=False,
        textfont=dict(size=fs_val, color=TXT_HI, family=F_NUM),
        customdata=list(df_data[y_col]),
        hovertemplate="<b>%{customdata}</b><br>%{x}<extra></extra>",
    ))
    fig.update_layout(**base_layout(height, dict(t=10, b=24, l=6, r=mr)))
    fig.update_layout(
        bargap=0.38,
        xaxis=dict(showticklabels=False, gridcolor=GRID, zeroline=False,
                   linecolor="rgba(0,0,0,0)",
                   range=[0, float(df_data[x_col].max()) * 1.18 if n else 1]),
        yaxis=dict(tickfont=dict(color=TXT_MID, size=fs_cat, family=F_TXT),
                   gridcolor="rgba(0,0,0,0)", linecolor="rgba(0,0,0,0)",
                   automargin=True, ticks=""),
    )
    return fig


def make_donut(df_data, name_col, val_col, height=380, money=True, centro_rot="Total"):
    """Rosca minimalista: percentual dentro da fatia, total no centro.

    Sem legenda lateral — a identificação fica na própria fatia e no hover,
    seguindo o mesmo padrão de "valor junto ao elemento" dos demais gráficos.
    """
    n = len(df_data)
    cores = [C_CRIT, C_WARN, C_QTY, C_INFO, C_OK, C_NEUT, "#22d3ee", "#fb923c",
             "#a78bfa", "#f472b6"]
    total = float(df_data[val_col].sum()) if n else 0
    txt_centro = fmt_brl0(total) if money else f"{int(total)}"

    fig = go.Figure(go.Pie(
        labels=[trunc(v, 22) for v in df_data[name_col]],
        values=df_data[val_col], hole=0.66, sort=False, direction="clockwise",
        marker=dict(colors=cores[:n] if n <= len(cores) else None,
                    line=dict(color="rgba(4,8,15,0.95)", width=2)),
        textinfo="percent", textposition="inside", insidetextorientation="horizontal",
        textfont=dict(size=13, color="#06101c", family=F_NUM),
        customdata=list(df_data[name_col]),
        hovertemplate="<b>%{customdata}</b><br>%{value:,.0f}<br>%{percent}<extra></extra>",
    ))
    fig.update_layout(**base_layout(height, dict(t=14, b=14, l=10, r=10)))
    fig.update_layout(annotations=[dict(
        text=(f"<span style='font-family:{F_NUM};font-size:19px;color:{TXT_HI}'>"
              f"<b>{txt_centro}</b></span><br>"
              f"<span style='font-size:11px;color:{TXT_LOW};letter-spacing:1.5px'>"
              f"{centro_rot.upper()}</span>"),
        x=0.5, y=0.5, showarrow=False, align="center")])
    return fig


def kpi(icon, label, value, sub, accent, delta=None):
    """delta: (texto, 'up'|'down'|'flat') — exibido apenas quando disponível."""
    d = ""
    if delta:
        d = f'<span class="kpi-delta kpi-{delta[1]}">{delta[0]}</span>'
    return (f'<div class="kpi" style="--acc:{accent};">'
            f'<div class="kpi-top"><div class="kpi-ico">{icon}</div>{d}</div>'
            f'<div class="kpi-val">{value}</div>'
            f'<div class="kpi-lab">{label}</div>'
            f'<div class="kpi-sub">{sub}</div></div>')


def agrupa_equipe(df_in, col_nome):
    """Agrupa a base filtrada por motorista ou entregador.

    Mantém apenas nomes preenchidos e devolve valor, quantidade de notas,
    clientes únicos e as placas em que a pessoa aparece — todos calculados
    sobre a mesma base já filtrada.
    """
    if col_nome not in df_in.columns or df_in.empty:
        return pd.DataFrame()
    base = df_in[df_in[col_nome].str.strip() != ""].copy()
    if base.empty:
        return pd.DataFrame()
    aggs = dict(Valor=(VALOR_COL, "sum"), Qtd=(VALOR_COL, "count"))
    if COL_CLIENTE:
        aggs["Clientes"] = (COL_CLIENTE, "nunique")
    out = base.groupby(col_nome).agg(**aggs).reset_index()
    if COL_PLACA:
        placas = (base.groupby(col_nome)[COL_PLACA]
                  .apply(lambda s: ", ".join(sorted(set([p for p in s if str(p).strip() != ""]))))
                  .reset_index().rename(columns={COL_PLACA: "Placas"}))
        out = out.merge(placas, on=col_nome, how="left")
    else:
        out["Placas"] = ""
    return out.sort_values("Valor", ascending=False)


# ═════════════════════════════════════════════════════════════════════════════
# PÁGINA: DASHBOARD
# ═════════════════════════════════════════════════════════════════════════════
if pagina == "Dashboard":

    # Indicador de evolução: mesma comparação já usada na seção semanal
    delta_val = None
    if usar_data and dt_sel:
        _prev = dt_sel - pd.Timedelta(days=7)
        if hasattr(_prev, "date"):
            _prev = _prev.date()
        _tot_prev = df_raw[df_raw["_DTENTREGA_DT"].dt.date == _prev][VALOR_COL].sum()
        if _tot_prev > 0:
            _p = (total_val - _tot_prev) / _tot_prev * 100
            delta_val = (f"{_p:+.1f}%", "up" if _p > 0 else "down" if _p < 0 else "flat")

    st.markdown(
        '<div class="kpi-grid">'
        + kpi("💰", "Valor total devolvido", fmt_brl(total_val), "Soma de VLTOTAL no filtro", "#22d3ee", delta_val)
        + kpi("📄", "Devoluções", f"{total_notas}", "Notas no filtro atual", "#a78bfa")
        + kpi("👥", "Clientes únicos", f"{total_clientes}", "Clientes com devolução", "#34d399")
        + kpi("📈", "Ticket médio", fmt_brl(ticket_medio), "Valor por devolução", "#fbbf24")
        + kpi("🚚", "Veículos únicos", f"{total_placas}", "Placas envolvidas", "#fb923c")
        + '</div>', unsafe_allow_html=True)

    # ── Lista de pedidos com devolução no dia filtrado ──────────────────────
    # Usa a mesma base já filtrada (df). Nenhum cálculo novo além da soma
    # exibida no rodapé do painel.
    _tag_ped = (f"DTENT {dt_sel.strftime('%d/%m/%Y')}" if usar_data and dt_sel
                else "Todos os períodos")
    panel_open("Pedidos com devolução", tag=_tag_ped, icon="🧾")

    PEDIDOS_CAMPOS = [
        (COL_NUMPED, "PEDIDO"),
        (COL_NF_VENDA, "NOTA VENDA"),
        (COL_NOTA_DEV, "NOTA DEVOLUÇÃO"),
        (COL_DTENTREGA, "DTENT"),
        (COL_NUMCAR, "NUMCAR"),
        (COL_PLACA, "PLACA"),
        (COL_MOT_NOME, "MOTORISTA"),
        (COL_ENT_NOME, "ENTREGADOR"),
        (COL_CODCLI, "CODCLI"),
        (COL_CLIENTE, "CLIENTE"),
        (COL_DESTINO, "DESTINO"),
        (COL_MOTIVO, "MOTIVO"),
        (COL_VENDEDOR, "VENDEDOR"),
        (COL_DEVOLUCION, "DEVOLUCIONISTA"),
    ]
    _cols_ped = [(o, a) for o, a in PEDIDOS_CAMPOS if o is not None]

    if not _cols_ped:
        st.warning("Nenhuma coluna de identificação de pedido encontrada na planilha.")
    elif df.empty:
        st.info("Nenhuma devolução para os filtros atuais. Ajuste a data ou limpe os filtros.")
    else:
        bp1, bp2 = st.columns([3, 1.4], gap="medium")
        with bp1:
            busca_ped = st.text_input(
                "Buscar", placeholder="Nota, pedido, cliente, placa, motorista ou motivo",
                key="ped_busca", label_visibility="collapsed")
        with bp2:
            _ordem_opts = ["Maior valor", "Menor valor", "Nota de venda", "Cliente", "Motorista"]
            if COL_NUMPED:
                _ordem_opts.insert(2, "Pedido")
            ordem_ped = st.selectbox("Ordenar", _ordem_opts, key="ped_ordem",
                                     label_visibility="collapsed")

        df_ped = df[[o for o, _ in _cols_ped] + [VALOR_COL]].copy()
        df_ped.columns = [a for _, a in _cols_ped] + ["_VALOR"]

        if busca_ped.strip():
            _t = busca_ped.strip()
            _mask = pd.Series([False] * len(df_ped), index=df_ped.index)
            for _c in [c for c in df_ped.columns if c != "_VALOR"]:
                _mask = _mask | df_ped[_c].astype(str).str.contains(_t, case=False, na=False)
            df_ped = df_ped[_mask]

        if ordem_ped == "Maior valor":
            df_ped = df_ped.sort_values("_VALOR", ascending=False)
        elif ordem_ped == "Menor valor":
            df_ped = df_ped.sort_values("_VALOR", ascending=True)
        elif ordem_ped == "Pedido" and "PEDIDO" in df_ped.columns:
            df_ped = df_ped.sort_values("PEDIDO")
        elif ordem_ped == "Nota de venda" and "NOTA VENDA" in df_ped.columns:
            df_ped = df_ped.sort_values("NOTA VENDA")
        elif ordem_ped == "Cliente" and "CLIENTE" in df_ped.columns:
            df_ped = df_ped.sort_values("CLIENTE")
        elif ordem_ped == "Motorista" and "MOTORISTA" in df_ped.columns:
            df_ped = df_ped.sort_values("MOTORISTA")

        _val_ped = df_ped["_VALOR"].sum()
        _qtd_ped = len(df_ped)
        _col_id_ped = "PEDIDO" if "PEDIDO" in df_ped.columns else (
            "NOTA VENDA" if "NOTA VENDA" in df_ped.columns else None)
        _lab_id_ped = "Pedidos distintos" if _col_id_ped == "PEDIDO" else "Notas de venda distintas"
        _notas_ped = df_ped[_col_id_ped].nunique() if _col_id_ped else _qtd_ped

        if _qtd_ped == 0:
            st.warning("Nenhum pedido encontrado para a busca informada.")
        else:
            df_ped_view = df_ped.copy()
            df_ped_view["VALOR"] = df_ped_view["_VALOR"].apply(fmt_brl)
            df_ped_view = df_ped_view.drop(columns=["_VALOR"])
            html_table(df_ped_view, min_width=1400)
            if _qtd_ped > 500:
                st.caption(f"Exibindo as primeiras 500 de {_qtd_ped} linhas.")

            st.markdown(
                f'<div class="cc-foot" style="margin-top:12px;">'
                f'  <div class="cc-foot-i">'
                f'    <span class="cc-foot-lab">Linhas de devolução</span>'
                f'    <span class="cc-foot-val" style="color:#38bdf8;">{_qtd_ped}</span>'
                f'  </div>'
                f'  <div class="cc-foot-sep"></div>'
                f'  <div class="cc-foot-i">'
                f'    <span class="cc-foot-lab">{_lab_id_ped}</span>'
                f'    <span class="cc-foot-val" style="color:#a78bfa;">{_notas_ped}</span>'
                f'  </div>'
                f'  <div class="cc-foot-sep"></div>'
                f'  <div class="cc-foot-i">'
                f'    <span class="cc-foot-lab">Valor devolvido</span>'
                f'    <span class="cc-foot-val" style="color:#34d399;">{fmt_brl(_val_ped)}</span>'
                f'  </div>'
                f'</div>', unsafe_allow_html=True)

            _suf = dt_sel.strftime("%Y%m%d") if (usar_data and dt_sel) else "todos"
            st.markdown("<div style='margin-top:12px;'></div>", unsafe_allow_html=True)
            export_tabela(df_ped_view, f"pedidos_devolvidos_{_suf}", "dl_pedidos",
                          sheet_name="Pedidos devolvidos")
    panel_close()

    # ── ÁREA 1 — gráficos por placa (empilhados, largura total) ─────────────
    hoje = date.today()
    primeiro_dia_mes = hoje.replace(day=1)
    df_mes = pd.DataFrame()
    if "_DTENTREGA_DT" in df_raw.columns:
        df_mes = df_raw[
            (df_raw["_DTENTREGA_DT"].dt.date >= primeiro_dia_mes) &
            (df_raw["_DTENTREGA_DT"].dt.date <= hoje)
        ].copy()

    periodo = (f"DTENT: {dt_sel.strftime('%d/%m/%Y')}" if usar_data and dt_sel
               else "Todos os períodos")

    panel_open("Devoluções por placa — valor e quantidade de notas", tag=periodo, icon="🚚")
    df_placa = pd.DataFrame()
    if COL_PLACA:
        _aggs = dict(Valor=(VALOR_COL, "sum"), Qtd=(VALOR_COL, "count"))
        if COL_CLIENTE:
            # Clientes únicos por placa, a partir da mesma base já filtrada.
            _aggs["Clientes"] = (COL_CLIENTE, "nunique")
        # Motorista e entregador vêm da aba NOMES — como são fixos por placa,
        # basta pegar o primeiro valor do grupo.
        _aggs["Motorista"] = (COL_MOT_NOME, "first")
        _aggs["Entregador"] = (COL_ENT_NOME, "first")
        df_placa = (df[df[COL_PLACA].str.strip() != ""]
                    .groupby(COL_PLACA).agg(**_aggs)
                    .reset_index().sort_values("Valor", ascending=False))
        if not df_placa.empty:
            _cd_placa = df_placa[["Motorista", "Entregador"]].values
            _eh_placa = ("<br>🧑‍✈️ %{customdata[0]}"
                         "<br>📦 %{customdata[1]}")
            st.plotly_chart(
                make_combo_chart(df_placa, COL_PLACA, "Valor", "Qtd", "", "", ramp(len(df_placa)),
                                 customdata=_cd_placa, extra_hover=_eh_placa),
                use_container_width=True, key="pc_1")
            st.markdown(
                '<div style="display:flex;gap:24px;flex-wrap:wrap;font-size:0.86rem;color:#7c8ea8;'
                'margin-top:-10px;padding-left:6px;">'
                '<span style="color:#ef4444;">● Top 5 crítico</span><span style="color:#f97316;">● 6–10 atenção</span><span style="color:#38bdf8;">● Demais</span>'
                '<span style="color:#facc15;">● Linha: notas devolvidas</span>'
                '<span style="color:#a78bfa;">Passe o mouse na barra para ver motorista e entregador</span></div>',
                unsafe_allow_html=True)
        else:
            st.info("Nenhuma placa no filtro atual. Ajuste a data ou limpe os filtros.")
    else:
        st.warning("Coluna PLACA não encontrada na planilha.")
    panel_close()

    # ── NOVO: devoluções por motorista (vinculado pela aba NOMES) ───────────
    df_motorista = agrupa_equipe(df, COL_MOT_NOME)
    panel_open("Devoluções por motorista — valor e quantidade de notas",
               tag=periodo, icon="🧑‍✈️")
    if df_motorista.empty:
        st.info("Sem motoristas vinculados no filtro atual.")
    else:
        _cd_mot = df_motorista[["Placas"]].values
        st.plotly_chart(
            make_combo_chart(df_motorista, COL_MOT_NOME, "Valor", "Qtd", "", "",
                             ramp(len(df_motorista)),
                             linha_nome="Notas devolvidas", linha_cor=C_QTY,
                             linha_ponto=P_QTY, linha_texto=C_QTY,
                             linha_rotulo="Notas",
                             customdata=_cd_mot, extra_hover="<br>🚚 %{customdata[0]}"),
            use_container_width=True, key="pc_2")
        st.markdown(
            '<div style="display:flex;gap:24px;flex-wrap:wrap;font-size:0.86rem;color:#7c8ea8;'
            'margin-top:-10px;padding-left:6px;">'
            '<span style="color:#ef4444;">● Top 5 crítico</span><span style="color:#f97316;">● 6–10 atenção</span><span style="color:#38bdf8;">● Demais</span>'
            '<span style="color:#facc15;">● Linha: notas devolvidas</span>'
            '<span>Vínculo placa → motorista pela aba NOMES</span></div>',
            unsafe_allow_html=True)
    panel_close()

    # ── NOVO: devoluções por entregador ────────────────────────────────────
    df_entregador = agrupa_equipe(df, COL_ENT_NOME)
    panel_open("Devoluções por entregador — valor e quantidade de notas",
               tag=periodo, icon="📦")
    if df_entregador.empty:
        st.info("Sem entregadores vinculados no filtro atual.")
    else:
        _cd_ent = df_entregador[["Placas"]].values
        st.plotly_chart(
            make_combo_chart(df_entregador, COL_ENT_NOME, "Valor", "Qtd", "", "",
                             ramp(len(df_entregador), 5, 10, "#a78bfa"),
                             linha_nome="Notas devolvidas", linha_cor=C_QTY,
                             linha_ponto=P_QTY, linha_texto=C_QTY,
                             linha_rotulo="Notas",
                             customdata=_cd_ent, extra_hover="<br>🚚 %{customdata[0]}"),
            use_container_width=True, key="pc_3")
        st.markdown(
            '<div style="display:flex;gap:24px;flex-wrap:wrap;font-size:0.86rem;color:#7c8ea8;'
            'margin-top:-10px;padding-left:6px;">'
            '<span style="color:#ef4444;">● Top 5 crítico</span><span style="color:#f97316;">● 6–10 atenção</span><span style="color:#38bdf8;">● Demais</span>'
            '<span style="color:#facc15;">● Linha: notas devolvidas</span>'
            '<span>Vínculo placa → entregador pela aba NOMES</span></div>',
            unsafe_allow_html=True)
    panel_close()

    # ── Gráfico irmão: clientes únicos por placa ────────────────────────────
    if COL_PLACA and COL_CLIENTE and not df_placa.empty and "Clientes" in df_placa.columns:
        panel_open("Devoluções por placa — valor e clientes únicos", tag=periodo, icon="👥")
        st.plotly_chart(
            make_combo_chart(df_placa, COL_PLACA, "Valor", "Clientes", "", "", ramp(len(df_placa)),
                             linha_nome="Clientes por veículo", linha_cor=C_OK,
                             linha_ponto=P_OK, linha_texto=C_OK, linha_rotulo="Clientes",
                             customdata=df_placa[["Motorista", "Entregador"]].values,
                             extra_hover="<br>🧑‍✈️ %{customdata[0]}<br>📦 %{customdata[1]}"),
            use_container_width=True, key="pc_4")
        st.markdown(
            '<div style="display:flex;gap:24px;flex-wrap:wrap;font-size:0.86rem;color:#7c8ea8;'
            'margin-top:-10px;padding-left:6px;">'
            '<span style="color:#ef4444;">● Top 5 crítico</span><span style="color:#f97316;">● 6–10 atenção</span><span style="color:#38bdf8;">● Demais</span>'
            '<span style="color:#22c55e;">● Linha: clientes únicos atendidos pelo veículo</span></div>',
            unsafe_allow_html=True)
        panel_close()

    # ── Acumulado do mês ────────────────────────────────────────────────────
    panel_open("Acumulado do mês", tag=hoje.strftime("%m/%Y"), icon="📈")
    if not df_mes.empty:
        df_mes_dia = (df_mes.assign(_DIA=df_mes["_DTENTREGA_DT"].dt.date)
                      .groupby("_DIA").agg(Valor=(VALOR_COL, "sum")).reset_index()
                      .sort_values("_DIA"))
        df_mes_dia["Acumulado"] = df_mes_dia["Valor"].cumsum()
        total_mes = df_mes_dia["Acumulado"].iloc[-1] if len(df_mes_dia) > 0 else 0
        valor_hoje = df_mes_dia.loc[df_mes_dia["_DIA"] == hoje, "Valor"].sum()

        _n_ac = len(df_mes_dia)
        _fs_ac = 14 if _n_ac <= 14 else 12 if _n_ac <= 24 else 11
        _max_ac = float(df_mes_dia["Acumulado"].max()) if _n_ac else 1

        fig_acum = go.Figure()
        fig_acum.add_trace(go.Scatter(
            x=df_mes_dia["_DIA"], y=df_mes_dia["Acumulado"],
            mode="lines+markers+text", fill="tozeroy",
            line=dict(color=C_INFO, width=2.2, shape="spline"),
            marker=dict(color=P_INFO, size=8, line=dict(color=C_INFO, width=2)),
            fillcolor="rgba(56,189,248,0.09)",
            text=[fmt_brl0(v) for v in df_mes_dia["Acumulado"]],
            textposition="top center", cliponaxis=False,
            textfont=dict(color=TXT_HI, size=_fs_ac, family=F_NUM),
            hovertemplate="<b>%{x}</b><br>Acumulado: R$ %{y:,.0f}<extra></extra>",
        ))
        fig_acum.update_layout(**base_layout(420, dict(t=44, b=46, l=10, r=14)))
        fig_acum.update_layout(
            xaxis=dict(tickfont=dict(size=_fs_ac, color=TXT_MID, family=F_NUM),
                       showgrid=False, linecolor=AXIS, zeroline=False,
                       ticks="outside", ticklen=4, tickcolor=AXIS, automargin=True),
            yaxis=eixo_y_oculto(_max_ac * 1.22),
        )
        st.plotly_chart(fig_acum, use_container_width=True, key="pc_5")
        st.markdown(
            f'<p style="font-size:0.95rem;color:#9fb2c9;text-align:center;margin-top:-8px;">'
            f'Hoje <b class="num" style="color:#facc15;font-size:1.3rem;">{fmt_brl0(valor_hoje)}</b>'
            f' &nbsp;·&nbsp; Mês <b class="num" style="color:#38bdf8;font-size:1.3rem;">{fmt_brl0(total_mes)}</b></p>',
            unsafe_allow_html=True)
    else:
        st.info("Nenhum lançamento no mês corrente ainda.")
    panel_close()

    # ── Evolução diária ─────────────────────────────────────────────────────
    panel_open("Evolução diária", tag="Mês corrente", icon="📊")
    if not df_mes.empty:
        df_ev = (df_mes.assign(_DIA=df_mes["_DTENTREGA_DT"].dt.date)
                 .groupby("_DIA").agg(Valor=(VALOR_COL, "sum"), Qtd=(VALOR_COL, "count"))
                 .reset_index().sort_values("_DIA"))
        fig_ev = go.Figure()
        _max_val_ev = float(df_ev["Valor"].max()) if len(df_ev) else 1
        _max_qtd_ev = float(df_ev["Qtd"].max()) if len(df_ev) else 1
        _n_ev = len(df_ev)
        _fs_ev = 14 if _n_ev <= 14 else 12 if _n_ev <= 24 else 11
        fig_ev.add_trace(go.Bar(
            x=df_ev["_DIA"], y=df_ev["Valor"], name="Valor (R$)",
            marker=dict(color=C_INFO, opacity=0.92,
                        line=dict(color="rgba(255,255,255,0.10)", width=1)),
            width=0.55,
            text=[fmt_brl0(v) for v in df_ev["Valor"]],
            textposition="outside", cliponaxis=False,
            textfont=dict(color=TXT_HI, size=_fs_ev, family=F_NUM),
            hovertemplate="<b>%{x}</b><br>Valor: R$ %{y:,.0f}<extra></extra>"))
        fig_ev.add_trace(go.Scatter(
            x=df_ev["_DIA"], y=df_ev["Qtd"], name="Notas", yaxis="y2",
            mode="lines+markers", line=dict(color=C_QTY, width=2.2, shape="spline"),
            marker=dict(color=P_QTY, size=8, line=dict(color=C_QTY, width=2)),
            hovertemplate="<b>%{x}</b><br>Notas: %{y}<extra></extra>"))
        fig_ev.update_layout(**base_layout(420, dict(t=48, b=46, l=10, r=14), legenda=True))
        fig_ev.update_layout(
            bargap=0.46,
            xaxis=dict(tickfont=dict(size=_fs_ev, color=TXT_MID, family=F_NUM),
                       showgrid=False, linecolor=AXIS, zeroline=False,
                       ticks="outside", ticklen=4, tickcolor=AXIS, automargin=True),
            yaxis=eixo_y_oculto(_max_val_ev * 1.30),
            yaxis2=dict(overlaying="y", side="right", showgrid=False,
                        showticklabels=False, zeroline=False,
                        range=[0, _max_qtd_ev * 1.60]),
        )
        _plot_h_ev = 420 - 48 - 46
        _ref_ev = [float(v) / (_max_val_ev * 1.30) * _plot_h_ev + 20 if _max_val_ev > 0 else 20
                   for v in df_ev["Valor"]]
        anotar_linha(fig_ev, list(df_ev["_DIA"]), list(df_ev["Qtd"]),
                     ref_pxs=_ref_ev, plot_h=_plot_h_ev,
                     frac_scale=(1 / (_max_qtd_ev * 1.60)) if _max_qtd_ev > 0 else 0,
                     cor=C_QTY, size=_fs_ev, gap=28)
        st.plotly_chart(fig_ev, use_container_width=True, key="pc_6")
    else:
        st.info("Nenhum lançamento no mês corrente ainda.")
    panel_close()

    # ── ÁREA 2 — painéis de motivo e veículo ────────────────────────────────
    a1, a2 = st.columns(2, gap="medium")

    with a1:
        panel_open("Devoluções por motivo", tag="Valor", icon="❗")
        if COL_MOTIVO:
            df_m_top = (df[df[COL_MOTIVO].str.strip() != ""]
                        .groupby(COL_MOTIVO).agg(Valor=(VALOR_COL, "sum"))
                        .reset_index().sort_values("Valor", ascending=True).tail(8))
            if not df_m_top.empty:
                st.plotly_chart(make_hbar(df_m_top, "Valor", COL_MOTIVO, RED, 400), use_container_width=True, key="pc_7")
            else:
                st.info("Sem motivos no filtro atual.")
        panel_close()

    with a2:
        panel_open("Devoluções por veículo", tag="Quantidade", icon="🚚")
        if COL_PLACA:
            df_v_top = (df[df[COL_PLACA].str.strip() != ""]
                        .groupby(COL_PLACA).agg(Qtd=(VALOR_COL, "count"))
                        .reset_index().sort_values("Qtd", ascending=True).tail(8))
            if not df_v_top.empty:
                st.plotly_chart(make_hbar(df_v_top, "Qtd", COL_PLACA, BLUE, 400, money=False),
                                use_container_width=True, key="pc_8")
            else:
                st.info("Sem veículos no filtro atual.")
        panel_close()

    # ── Comparação semanal por placa (lógica original preservada) ───────────
    if COL_PLACA:
        dia_ref = dt_sel if (usar_data and dt_sel) else date.today()
        dia_sem_passada = dia_ref - pd.Timedelta(days=7)
        dias_semana_pt = ["Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira",
                          "Sexta-feira", "Sábado", "Domingo"]
        nome_dia = dias_semana_pt[dia_ref.weekday()]

        if "_DTENTREGA_DT" in df_raw.columns:
            df_atual_dia = df_raw[
                (df_raw["_DTENTREGA_DT"].dt.date == dia_ref) & (df_raw[COL_PLACA].str.strip() != "")
            ].groupby(COL_PLACA).agg(Valor=(VALOR_COL, "sum"), Qtd=(VALOR_COL, "count")).reset_index()
            df_semana_dia = df_raw[
                (df_raw["_DTENTREGA_DT"].dt.date == dia_sem_passada) & (df_raw[COL_PLACA].str.strip() != "")
            ].groupby(COL_PLACA).agg(Valor=(VALOR_COL, "sum"), Qtd=(VALOR_COL, "count")).reset_index()
        else:
            df_atual_dia = pd.DataFrame(columns=[COL_PLACA, "Valor", "Qtd"])
            df_semana_dia = pd.DataFrame(columns=[COL_PLACA, "Valor", "Qtd"])

        todas_placas_comp = sorted(set(df_atual_dia[COL_PLACA]))

        if todas_placas_comp:
            df_comp = pd.DataFrame({COL_PLACA: todas_placas_comp})
            df_comp = df_comp.merge(df_atual_dia.rename(columns={"Valor": "Valor_Atual", "Qtd": "Qtd_Atual"}),
                                    on=COL_PLACA, how="left")
            df_comp = df_comp.merge(df_semana_dia.rename(columns={"Valor": "Valor_Semana", "Qtd": "Qtd_Semana"}),
                                    on=COL_PLACA, how="left")
            df_comp[["Valor_Atual", "Qtd_Atual", "Valor_Semana", "Qtd_Semana"]] = \
                df_comp[["Valor_Atual", "Qtd_Atual", "Valor_Semana", "Qtd_Semana"]].fillna(0)
            df_comp = df_comp.sort_values("Valor_Atual", ascending=False)

            # Nome da equipe da placa, vindo da aba NOMES (apenas para o tooltip)
            df_comp["_MOT"] = df_comp[COL_PLACA].apply(
                lambda p: mapa_motorista.get(norm_placa(p), SEM_CADASTRO))
            df_comp["_ENT"] = df_comp[COL_PLACA].apply(
                lambda p: mapa_entregador.get(norm_placa(p), SEM_CADASTRO))

            # Totais e variação — mesmas fórmulas de antes, apenas calculadas
            # antes da renderização para alimentar o cabeçalho e o rodapé do card.
            total_atual = df_comp["Valor_Atual"].sum()
            total_semana = df_comp["Valor_Semana"].sum()
            var_pct = ((total_atual - total_semana) / total_semana * 100) if total_semana > 0 else 0

            _lbl_ref = dia_ref.strftime("%d/%m")
            _lbl_ant = dia_sem_passada.strftime("%d/%m")

            with st.container(border=True):
                # ── Cabeçalho do card ───────────────────────────────────────
                st.markdown(f"""
                <div class="cc-head">
                  <div class="cc-head-l">
                    <p class="cc-title">Devoluções por placa — valor e quantidade</p>
                    <p class="cc-sub">Comparativo: {_lbl_ref} (referência) vs {_lbl_ant} (semana passada) · {nome_dia}</p>
                  </div>
                  <div class="cc-head-r">
                    <div class="cc-mini">
                      <span class="cc-mini-lab">Total {_lbl_ref}</span>
                      <span class="cc-mini-val" style="color:#38bdf8;">{fmt_brl0(total_atual)}</span>
                      <span class="cc-mini-sub">valor devolvido</span>
                    </div>
                    <div class="cc-mini">
                      <span class="cc-mini-lab">Total {_lbl_ant}</span>
                      <span class="cc-mini-val" style="color:#cbd5e1;">{fmt_brl0(total_semana)}</span>
                      <span class="cc-mini-sub">valor devolvido</span>
                    </div>
                    <div class="cc-ctrls">
                      <span class="cc-pick">Por placa ▾</span>
                      <span class="cc-dots">⋮</span>
                    </div>
                  </div>
                </div>
                """, unsafe_allow_html=True)

                # ── Tooltip único por placa (mesmos números do gráfico) ─────
                cdata = [[fmt_brl0(va), fmt_brl0(vs), int(qa), int(qs), mt, en]
                         for va, vs, qa, qs, mt, en in zip(df_comp["Valor_Atual"], df_comp["Valor_Semana"],
                                                           df_comp["Qtd_Atual"], df_comp["Qtd_Semana"],
                                                           df_comp["_MOT"], df_comp["_ENT"])]
                htmpl = (f"<b>Placa %{{x}}</b><br>"
                         f"🧑‍✈️ %{{customdata[4]}}<br>"
                         f"📦 %{{customdata[5]}}<br>"
                         f"<span style='color:#38bdf8'>●</span> {_lbl_ref} — %{{customdata[0]}}<br>"
                         f"<span style='color:#cbd5e1'>●</span> {_lbl_ant} — %{{customdata[1]}}<br>"
                         f"<span style='color:#fbbf24'>●</span> Notas {_lbl_ref} — %{{customdata[2]}}<br>"
                         f"<span style='color:#8fa3bb'>●</span> Notas {_lbl_ant} — %{{customdata[3]}}"
                         f"<extra></extra>")

                n_comp = len(df_comp)
                max_qtd_comp = max(df_comp["Qtd_Atual"].max(), df_comp["Qtd_Semana"].max(), 1)
                max_val_comp = max(df_comp["Valor_Atual"].max(), df_comp["Valor_Semana"].max(), 1)

                # Números do comparativo: legibilidade tem prioridade sobre a
                # densidade — mesmo com muitas placas o valor continua grande.
                _fs_v_comp = 19 if n_comp <= 10 else 17 if n_comp <= 16 else 15
                _fs_c_comp = 15 if n_comp <= 10 else 14 if n_comp <= 18 else 12
                # Rótulo em branco só onde há valor: barra zerada polui a leitura.
                _txt_at = [fmt_brl0(v) if v > 0 else "" for v in df_comp["Valor_Atual"]]
                _txt_sm = [fmt_brl0(v) if v > 0 else "" for v in df_comp["Valor_Semana"]]

                fig_comp = go.Figure()
                # Série histórica em cinza neutro, série de referência em azul:
                # a cor diz o papel do dado, não o gráfico.
                fig_comp.add_trace(go.Bar(
                    x=df_comp[COL_PLACA], y=df_comp["Valor_Semana"],
                    name=f"{_lbl_ant} — semana passada",
                    marker=dict(color="rgba(143,163,187,0.38)", opacity=1,
                                line=dict(color="rgba(203,213,225,0.35)", width=1)),
                    width=0.30,
                    text=_txt_sm, textposition="outside", cliponaxis=False,
                    textfont=dict(size=_fs_v_comp, color=P_NEUT, family=F_NUM),
                    customdata=cdata, hovertemplate=htmpl))
                fig_comp.add_trace(go.Bar(
                    x=df_comp[COL_PLACA], y=df_comp["Valor_Atual"],
                    name=f"{_lbl_ref} — referência",
                    marker=dict(color=C_INFO, opacity=0.95,
                                line=dict(color="rgba(255,255,255,0.12)", width=1)),
                    width=0.30,
                    text=_txt_at, textposition="outside", cliponaxis=False,
                    textfont=dict(size=_fs_v_comp + 2, color=TXT_HI, family=F_NUM),
                    customdata=cdata, hovertemplate=htmpl))
                fig_comp.add_trace(go.Scatter(
                    x=df_comp[COL_PLACA], y=df_comp["Qtd_Semana"],
                    name="Notas — semana passada", mode="lines+markers",
                    line=dict(color=C_NEUT, width=1.8, dash="dot", shape="spline"),
                    marker=dict(color=P_NEUT, size=9, line=dict(color="rgba(4,8,15,0.9)", width=1.8)),
                    customdata=cdata, hovertemplate=htmpl, yaxis="y2"))
                fig_comp.add_trace(go.Scatter(
                    x=df_comp[COL_PLACA], y=df_comp["Qtd_Atual"],
                    name="Notas — referência", mode="lines+markers",
                    line=dict(color=C_QTY, width=2.8, shape="spline"),
                    marker=dict(color=P_QTY, size=11, line=dict(color="rgba(4,8,15,0.9)", width=2)),
                    customdata=cdata, hovertemplate=htmpl, yaxis="y2"))

                _h_comp = max(540, min(n_comp * 66, 820))
                fig_comp.update_layout(**base_layout(_h_comp, dict(t=70, b=104, l=8, r=12),
                                                     legenda=True))
                fig_comp.update_layout(
                    barmode="group", bargap=0.44, bargroupgap=0.14,
                    xaxis=eixo_x(df_comp[COL_PLACA], size=_fs_c_comp),
                    yaxis=eixo_y_oculto(max_val_comp * 1.52),
                    yaxis2=dict(overlaying="y", side="right", showgrid=False, zeroline=False,
                                showticklabels=False, range=[0, max_qtd_comp * 3.6]),
                )
                # ── Rótulos das duas linhas, posicionados sem colisão ───────
                _plot_h = _h_comp - 70 - 104
                _ref_barras = [
                    [float(va) / (max_val_comp * 1.52) * _plot_h + 20,
                     float(vs) / (max_val_comp * 1.52) * _plot_h + 20]
                    for va, vs in zip(df_comp["Valor_Atual"], df_comp["Valor_Semana"])]
                _fs = (1 / (max_qtd_comp * 3.6)) if max_qtd_comp > 0 else 0
                _ocup = anotar_linha(fig_comp, list(df_comp[COL_PLACA]), list(df_comp["Qtd_Atual"]),
                                     ref_pxs=_ref_barras, plot_h=_plot_h, frac_scale=_fs,
                                     cor=C_QTY, size=_fs_v_comp, gap=34)
                anotar_linha(fig_comp, list(df_comp[COL_PLACA]), list(df_comp["Qtd_Semana"]),
                             ref_pxs=_ref_barras, plot_h=_plot_h, frac_scale=_fs,
                             cor=C_NEUT, size=_fs_v_comp, ocupados=_ocup, gap=34)

                st.plotly_chart(fig_comp, use_container_width=True,
                                config={"displayModeBar": False}, key="pc_9")

                # ── Rodapé do card ──────────────────────────────────────────
                _var_cor = C_CRIT if var_pct > 0 else C_OK if var_pct < 0 else C_NEUT
                _var_seta = "▲" if var_pct > 0 else "▼" if var_pct < 0 else "■"
                st.markdown(f"""
                <div class="cc-foot">
                  <div class="cc-foot-i">
                    <span class="cc-foot-lab"><i class="d" style="background:#38bdf8;"></i>{_lbl_ref} — referência</span>
                    <span class="cc-foot-val" style="color:#38bdf8;">{fmt_brl0(total_atual)}</span>
                  </div>
                  <div class="cc-foot-sep"></div>
                  <div class="cc-foot-i">
                    <span class="cc-foot-lab"><i class="d" style="background:#8fa3bb;"></i>{_lbl_ant} — semana passada</span>
                    <span class="cc-foot-val" style="color:#cbd5e1;">{fmt_brl0(total_semana)}</span>
                  </div>
                  <div class="cc-foot-sep"></div>
                  <div class="cc-foot-i">
                    <span class="cc-foot-lab">Variação</span>
                    <span class="cc-foot-val" style="color:{_var_cor};">{_var_seta} {var_pct:+.1f}%</span>
                  </div>
                </div>
                """, unsafe_allow_html=True)
        else:
            panel_open(f"Comparação por placa — {nome_dia}",
                       tag=f"{dia_ref.strftime('%d/%m')} × {dia_sem_passada.strftime('%d/%m')}", icon="📆")
            st.info("Sem registros de placa nas datas comparadas.")
            panel_close()

    # ── Ocorrências de retorno ──────────────────────────────────────────────
    if COL_MOTIVO:
        panel_open("Ocorrências de retorno", tag="Valor e quantidade", icon="❗")
        df_mot_v = (df[df[COL_MOTIVO].str.strip() != ""]
                    .groupby(COL_MOTIVO).agg(Valor=(VALOR_COL, "sum"), Qtd=(VALOR_COL, "count"))
                    .reset_index().sort_values("Valor", ascending=False))
        if not df_mot_v.empty:
            n_m = len(df_mot_v)
            fig_mv = make_combo_chart(df_mot_v, COL_MOTIVO, "Valor", "Qtd", "", "", ramp(n_m, 3, 6))
            fig_mv.update_layout(height=max(430, min(n_m * 58, 660)), margin=dict(t=54, b=110, l=10, r=30))
            fig_mv.update_xaxes(tickangle=-35, automargin=True)
            st.plotly_chart(fig_mv, use_container_width=True, key="pc_10")
            with st.expander("Ver tabela de motivos"):
                df_mt = df_mot_v.copy()
                df_mt["Valor (R$)"] = df_mt["Valor"].apply(fmt_brl)
                df_mt["% Total"] = (df_mt["Valor"] / total_val * 100).round(1).astype(str) + "%" if total_val > 0 else "0%"
                df_mt = df_mt.rename(columns={COL_MOTIVO: "Motivo", "Qtd": "Qtd."})
                html_table(df_mt[["Motivo", "Qtd.", "Valor (R$)", "% Total"]], max_rows=200, min_width=600)
        else:
            st.info("Sem ocorrências no filtro atual.")
        panel_close()

    # ── Rankings ────────────────────────────────────────────────────────────
    c1, c2, c3 = st.columns(3, gap="medium")
    with c1:
        panel_open("Top motivos", tag="R$", icon="❗")
        if COL_MOTIVO:
            df_m2 = (df[df[COL_MOTIVO].str.strip() != ""]
                     .groupby(COL_MOTIVO).agg(Valor=(VALOR_COL, "sum"), Qtd=(VALOR_COL, "count"))
                     .reset_index().sort_values("Valor", ascending=True).tail(8))
            if not df_m2.empty:
                st.plotly_chart(make_hbar(df_m2, "Valor", COL_MOTIVO, RED, 400), use_container_width=True, key="pc_11")
                top = df_m2.iloc[-1]
                pct = top["Valor"] / total_val * 100 if total_val > 0 else 0
                st.markdown(f'<p style="font-size:0.73rem;color:#7c8ea8;padding:0 4px 10px;">'
                            f'{top[COL_MOTIVO]} — {pct:.1f}% ({fmt_brl0(top["Valor"])})</p>',
                            unsafe_allow_html=True)
        panel_close()
    with c2:
        panel_open("Top 10 clientes", tag="R$", icon="👥")
        if COL_CLIENTE:
            df_cl = (df[df[COL_CLIENTE].str.strip() != ""]
                     .groupby(COL_CLIENTE).agg(Valor=(VALOR_COL, "sum"), Qtd=(VALOR_COL, "count"))
                     .reset_index().sort_values("Valor", ascending=True).tail(10))
            if not df_cl.empty:
                st.plotly_chart(make_hbar(df_cl, "Valor", COL_CLIENTE, MIXED, 400), use_container_width=True, key="pc_12")
                top_c = df_cl.iloc[-1]
                st.markdown(f'<p style="font-size:0.73rem;color:#7c8ea8;padding:0 4px 10px;">'
                            f'{str(top_c[COL_CLIENTE])[:30]} — {fmt_brl0(top_c["Valor"])}</p>',
                            unsafe_allow_html=True)
        panel_close()
    with c3:
        panel_open("Top 10 vendedores", tag="NOMERCA", icon="🧑‍💼")
        if COL_VENDEDOR:
            df_vv = (df[df[COL_VENDEDOR].str.strip() != ""]
                     .groupby(COL_VENDEDOR).agg(Valor=(VALOR_COL, "sum"), Qtd=(VALOR_COL, "count"))
                     .reset_index().sort_values("Valor", ascending=True).tail(10))
            if not df_vv.empty:
                st.plotly_chart(make_hbar(df_vv, "Valor", COL_VENDEDOR, BLUE, 400), use_container_width=True, key="pc_13")
                top_v = df_vv.iloc[-1]
                st.markdown(f'<p style="font-size:0.73rem;color:#7c8ea8;padding:0 4px 10px;">'
                            f'{str(top_v[COL_VENDEDOR])[:30]} — {int(top_v["Qtd"])} devoluções</p>',
                            unsafe_allow_html=True)
        panel_close()

    c4, c5 = st.columns([1, 2], gap="medium")
    with c4:
        panel_open("Por destino", tag="Top 10", icon="🏙️")
        if COL_DESTINO:
            df_dd = (df[df[COL_DESTINO].str.strip() != ""]
                     .groupby(COL_DESTINO).agg(Valor=(VALOR_COL, "sum"))
                     .reset_index().sort_values("Valor", ascending=False).head(10))
            if not df_dd.empty:
                fig_dd = make_donut(df_dd, COL_DESTINO, "Valor", height=380,
                                    money=True, centro_rot="Devolvido")
                st.plotly_chart(fig_dd, use_container_width=True, key="pc_14")
        panel_close()
    with c5:
        panel_open("Ranking de motivos", tag="Consolidado", icon="📊")
        if COL_MOTIVO:
            df_rk = (df.groupby(COL_MOTIVO).agg(Qtd=(VALOR_COL, "count"), Total=(VALOR_COL, "sum"))
                     .reset_index().sort_values("Total", ascending=False))
            df_rk["Valor Total"] = df_rk["Total"].apply(fmt_brl)
            df_rk["% Total"] = (df_rk["Total"] / total_val * 100).round(1).astype(str) + "%" if total_val > 0 else "0%"
            html_table(df_rk.rename(columns={COL_MOTIVO: "Motivo"})[["Motivo", "Qtd", "Valor Total", "% Total"]],
                       max_rows=200, min_width=520)
        panel_close()


# ═════════════════════════════════════════════════════════════════════════════
# PÁGINA: EQUIPE (motoristas e entregadores — vínculo pela aba NOMES)
# ═════════════════════════════════════════════════════════════════════════════
elif pagina == "Equipe":
    df_mot_pg = agrupa_equipe(df, COL_MOT_NOME)
    df_ent_pg = agrupa_equipe(df, COL_ENT_NOME)

    st.markdown(
        '<div class="kpi-grid">'
        + kpi("🧑‍✈️", "Motoristas", f"{total_motoristas}", "Com devolução no filtro", "#fbbf24")
        + kpi("📦", "Entregadores", f"{total_entregadores}", "Com devolução no filtro", "#a78bfa")
        + kpi("💰", "Valor total", fmt_brl(total_val), "Soma de VLTOTAL", "#22d3ee")
        + kpi("📄", "Devoluções", f"{total_notas}", "Notas no filtro", "#34d399")
        + kpi("🚚", "Placas cadastradas", f"{_n_nomes}", "Linhas na aba NOMES", "#fb923c")
        + '</div>', unsafe_allow_html=True)

    if nomes_load_error:
        st.error(f"A aba NOMES não pôde ser lida: {nomes_load_error}")
        st.markdown("""
**Como resolver**
1. Abra a planilha no Google Sheets.
2. Confirme que a aba se chama exatamente **NOMES**.
3. Vá em **Compartilhar → Acesso geral → Qualquer pessoa com o link (leitor)**
   ou publique a aba em **Arquivo → Compartilhar → Publicar na web**.
4. Volte aqui e clique em **Atualizar dados**.
        """)

    panel_open("Motoristas — valor e quantidade de notas",
               tag=f"{len(df_mot_pg)} motoristas", icon="🧑‍✈️")
    if df_mot_pg.empty:
        st.info("Sem motoristas no filtro atual.")
    else:
        fig_mp = make_combo_chart(df_mot_pg, COL_MOT_NOME, "Valor", "Qtd", "", "",
                                  ramp(len(df_mot_pg)),
                                  customdata=df_mot_pg[["Placas"]].values,
                                  extra_hover="<br>🚚 %{customdata[0]}")
        fig_mp.update_xaxes(tickangle=-35, automargin=True)
        st.plotly_chart(fig_mp, use_container_width=True, key="pc_15")
    panel_close()

    panel_open("Entregadores — valor e quantidade de notas",
               tag=f"{len(df_ent_pg)} entregadores", icon="📦")
    if df_ent_pg.empty:
        st.info("Sem entregadores no filtro atual.")
    else:
        fig_ep = make_combo_chart(df_ent_pg, COL_ENT_NOME, "Valor", "Qtd", "", "",
                                  ramp(len(df_ent_pg), 5, 10, "#a78bfa"),
                                  linha_cor=C_QTY, linha_ponto=P_QTY,
                                  linha_texto=C_QTY,
                                  customdata=df_ent_pg[["Placas"]].values,
                                  extra_hover="<br>🚚 %{customdata[0]}")
        fig_ep.update_xaxes(tickangle=-35, automargin=True)
        st.plotly_chart(fig_ep, use_container_width=True, key="pc_16")
    panel_close()

    q1, q2 = st.columns(2, gap="medium")
    with q1:
        panel_open("Detalhamento por motorista", tag="Consolidado", icon="📋")
        if not df_mot_pg.empty:
            tm = df_mot_pg.copy()
            tm["Valor total"] = tm["Valor"].apply(fmt_brl)
            tm["Ticket médio"] = (tm["Valor"] / tm["Qtd"]).apply(fmt_brl)
            tm["% do total"] = (tm["Valor"] / total_val * 100).round(1).astype(str) + "%" if total_val > 0 else "0%"
            tm = tm.rename(columns={COL_MOT_NOME: "Motorista", "Qtd": "Notas"})
            _cols_tm = ["Motorista", "Placas", "Notas", "Valor total", "Ticket médio", "% do total"]
            html_table(tm[[c for c in _cols_tm if c in tm.columns]], min_width=820)
            st.markdown("<div style='margin-top:12px;'></div>", unsafe_allow_html=True)
            export_tabela(tm[[c for c in _cols_tm if c in tm.columns]], "motoristas",
                          "dl_mot", sheet_name="Motoristas")
        else:
            st.info("Sem dados para listar.")
        panel_close()
    with q2:
        panel_open("Detalhamento por entregador", tag="Consolidado", icon="📋")
        if not df_ent_pg.empty:
            te = df_ent_pg.copy()
            te["Valor total"] = te["Valor"].apply(fmt_brl)
            te["Ticket médio"] = (te["Valor"] / te["Qtd"]).apply(fmt_brl)
            te["% do total"] = (te["Valor"] / total_val * 100).round(1).astype(str) + "%" if total_val > 0 else "0%"
            te = te.rename(columns={COL_ENT_NOME: "Entregador", "Qtd": "Notas"})
            _cols_te = ["Entregador", "Placas", "Notas", "Valor total", "Ticket médio", "% do total"]
            html_table(te[[c for c in _cols_te if c in te.columns]], min_width=820)
            st.markdown("<div style='margin-top:12px;'></div>", unsafe_allow_html=True)
            export_tabela(te[[c for c in _cols_te if c in te.columns]], "entregadores",
                          "dl_ent", sheet_name="Entregadores")
        else:
            st.info("Sem dados para listar.")
        panel_close()

    panel_open("Cadastro da aba NOMES", tag=f"{_n_nomes} placas", icon="🗂️")
    if df_nomes_norm is not None and not df_nomes_norm.empty:
        _tb_nomes = df_nomes_norm.drop(columns=[c for c in ["_PLACA_KEY"] if c in df_nomes_norm.columns])
        html_table(_tb_nomes, min_width=760)
        st.markdown("<div style='margin-top:12px;'></div>", unsafe_allow_html=True)
        export_tabela(_tb_nomes, "cadastro_nomes", "dl_nomes", sheet_name="NOMES")
        if placas_sem_cadastro:
            st.caption(f"Placas presentes nas devoluções e ausentes no cadastro "
                       f"({len(placas_sem_cadastro)}): {', '.join(placas_sem_cadastro[:25])}"
                       f"{'…' if len(placas_sem_cadastro) > 25 else ''}")
        else:
            st.caption("Todas as placas da base de devoluções estão cadastradas na aba NOMES.")
    else:
        st.info("Cadastro de nomes indisponível nesta sessão.")
    panel_close()


# ═════════════════════════════════════════════════════════════════════════════
# PÁGINA: REENTREGAS
# ═════════════════════════════════════════════════════════════════════════════
elif pagina == "Reentregas":
    if reent_load_error:
        st.error(f"Não foi possível carregar as reentregas: {reent_load_error}")
        st.markdown("""
**Como resolver**
1. Abra a planilha no Google Sheets.
2. Vá em **Arquivo → Compartilhar → Publicar na web**.
3. Selecione a aba **8261 - REENTREGAS 2026**.
4. Clique em **Publicar** e confirme.
5. Volte aqui e use **Atualizar dados**.
        """)
        with st.expander("URLs tentadas"):
            for u in REENTREGAS_URLS:
                st.code(u)
    elif df_reent is None or len(df_reent) == 0:
        st.warning("Nenhum dado encontrado na aba de reentregas.")
    else:
        st.markdown('<div class="filters"><div class="filters-h"><span class="pip"></span>'
                    'Filtros — reentregas</div>', unsafe_allow_html=True)
        rf1, rf2, rf3, rf4 = st.columns([3, 3, 1.3, 1], gap="medium")
        usar_data_reent = False
        dt_reent_sel = None

        with rf1:
            datas_reent_ok = df_reent["_DATATRANSF_DT"].dropna()
            dt_col_reent = reent_cols.get("DATATRANSF")
            col_label_reent = dt_col_reent if dt_col_reent else "DTRANSF"
            if len(datas_reent_ok) > 0:
                datas_reent_unicas = sorted(datas_reent_ok.dt.date.unique())
                opcoes_reent = ["— Todas as datas —"] + [d.strftime("%d/%m/%Y") for d in datas_reent_unicas]
                sel_reent_str = st.selectbox(f"Data ({col_label_reent})", opcoes_reent, key="r_dtsel")
                if sel_reent_str != "— Todas as datas —":
                    dt_reent_sel = datetime.strptime(sel_reent_str, "%d/%m/%Y").date()
                    usar_data_reent = True
            else:
                st.caption("Sem datas DTRANSF válidas.")

        with rf2:
            motivo_reent_col = reent_cols.get("MOTIVOTRANSF")
            if motivo_reent_col:
                mot_reent_opts = sorted([x for x in df_reent[motivo_reent_col].unique()
                                         if x not in ("", "N/D", "nan", "None")])
                sel_mot_reent = st.multiselect("Motivo da transferência", mot_reent_opts,
                                               default=[], key="r_mot", placeholder="Todos")
            else:
                sel_mot_reent = []

        with rf3:
            st.markdown("<div style='height:26px'></div>", unsafe_allow_html=True)
            if st.button("Atualizar dados", use_container_width=True, type="primary", key="btn_reent"):
                st.cache_data.clear()
                st.rerun()
        with rf4:
            st.markdown("<div style='height:26px'></div>", unsafe_allow_html=True)
            if st.button("Limpar", use_container_width=True, key="btn_clear_reent"):
                for k in ("r_dtsel", "r_mot"):
                    st.session_state.pop(k, None)
                st.rerun()

        st.markdown('</div>', unsafe_allow_html=True)

        df_r = df_reent.copy()
        if usar_data_reent and dt_reent_sel:
            df_r = df_r[df_r["_DATATRANSF_DT"].dt.date == dt_reent_sel]
        if sel_mot_reent and motivo_reent_col:
            df_r = df_r[df_r[motivo_reent_col].isin(sel_mot_reent)]

        if usar_data_reent and dt_reent_sel:
            st.info(f"Data {dt_reent_sel.strftime('%d/%m/%Y')} — **{len(df_r)} registros**")

        total_reent = len(df_r)
        total_reent_valor = df_r[REENT_VALOR_COL].sum() if REENT_VALOR_COL in df_r.columns else 0
        placaant_col = reent_cols.get("PLACAANT")
        cliente_r_col = reent_cols.get("CLIENTE")
        praca_r_col = reent_cols.get("PRACA")
        nome_r_col = reent_cols.get("NOME")
        motivo_r_col2 = reent_cols.get("MOTIVOTRANSF")
        total_placas_r = df_r[placaant_col].nunique() if placaant_col and placaant_col in df_r.columns else 0
        total_cli_r = df_r[cliente_r_col].nunique() if cliente_r_col and cliente_r_col in df_r.columns else 0
        ticket_r = total_reent_valor / total_reent if total_reent > 0 and total_reent_valor > 0 else 0
        total_pracas_r = df_r[praca_r_col].nunique() if praca_r_col and praca_r_col in df_r.columns else 0

        if total_reent_valor > 0:
            st.markdown(
                '<div class="kpi-grid">'
                + kpi("🔄", "Total de reentregas", f"{total_reent}", "Registros no filtro", "#fbbf24")
                + kpi("💰", "Valor total", fmt_brl(total_reent_valor), "Soma de VLTOTGER", "#22d3ee")
                + kpi("👥", "Clientes únicos", f"{total_cli_r}", "Clientes atendidos", "#34d399")
                + kpi("📈", "Ticket médio", fmt_brl(ticket_r), "Valor por reentrega", "#a78bfa")
                + kpi("🚚", "Placas anteriores", f"{total_placas_r}", "PLACAANT distintas", "#fb923c")
                + '</div>', unsafe_allow_html=True)
        else:
            st.markdown(
                '<div class="kpi-grid">'
                + kpi("🔄", "Total de reentregas", f"{total_reent}", "Registros no filtro", "#fbbf24")
                + kpi("👥", "Clientes únicos", f"{total_cli_r}", "Clientes atendidos", "#34d399")
                + kpi("🚚", "Placas anteriores", f"{total_placas_r}", "PLACAANT distintas", "#fb923c")
                + kpi("🗺️", "Praças", f"{total_pracas_r}", "Praças atendidas", "#a78bfa")
                + '</div>', unsafe_allow_html=True)

        g1, g2 = st.columns(2, gap="medium")
        with g1:
            panel_open("Reentregas por placa anterior", tag="PLACAANT", icon="🚚")
            if placaant_col and placaant_col in df_r.columns:
                df_pr = (df_r[df_r[placaant_col].str.strip() != ""]
                         .groupby(placaant_col).agg(Qtd=(placaant_col, "count"))
                         .reset_index().sort_values("Qtd", ascending=False))
                if not df_pr.empty:
                    st.plotly_chart(make_bar_simple(df_pr, placaant_col, "Qtd", ramp(len(df_pr), 3, 6, C_OK)),
                                    use_container_width=True, key="pc_17")
                else:
                    st.info("Sem placas no filtro atual.")
            else:
                st.warning("Coluna PLACAANT não encontrada.")
            panel_close()

        with g2:
            panel_open("Reentregas por motivo", tag="MOTIVOTRANSF", icon="❗")
            if motivo_r_col2 and motivo_r_col2 in df_r.columns:
                df_mr = (df_r[df_r[motivo_r_col2].str.strip() != ""]
                         .groupby(motivo_r_col2).agg(Qtd=(motivo_r_col2, "count"))
                         .reset_index().sort_values("Qtd", ascending=False))
                if not df_mr.empty:
                    st.plotly_chart(make_bar_simple(df_mr, motivo_r_col2, "Qtd", ramp(len(df_mr), 3, 6)),
                                    use_container_width=True, key="pc_18")
                else:
                    st.info("Sem motivos no filtro atual.")
            else:
                st.warning("Coluna MOTIVOTRANSF não encontrada.")
            panel_close()

        # ── Reentregas por motorista da placa anterior (aba NOMES) ──────────
        if placaant_col and placaant_col in df_r.columns and mapa_motorista:
            panel_open("Reentregas por motorista da placa anterior", tag="Aba NOMES", icon="🧑‍✈️")
            _dfr_nome = df_r[df_r[placaant_col].str.strip() != ""].copy()
            _dfr_nome["_MOT"] = _dfr_nome[placaant_col].apply(
                lambda p: str(mapa_motorista.get(norm_placa(p), SEM_CADASTRO)).upper())
            df_mot_r = (_dfr_nome.groupby("_MOT").agg(Qtd=("_MOT", "count"))
                        .reset_index().sort_values("Qtd", ascending=False))
            if not df_mot_r.empty:
                st.plotly_chart(make_bar_simple(df_mot_r, "_MOT", "Qtd", ramp(len(df_mot_r), 3, 6)),
                                use_container_width=True, key="pc_19")
            else:
                st.info("Sem motoristas vinculados no filtro atual.")
            panel_close()

        cr1, cr2, cr3 = st.columns(3, gap="medium")
        with cr1:
            panel_open("Principais motivos", tag="Qtd", icon="❗")
            if motivo_r_col2 and motivo_r_col2 in df_r.columns:
                df_mr2 = (df_r[df_r[motivo_r_col2].str.strip() != ""]
                          .groupby(motivo_r_col2).agg(Qtd=(motivo_r_col2, "count"))
                          .reset_index().sort_values("Qtd", ascending=True).tail(8))
                if not df_mr2.empty:
                    st.plotly_chart(make_hbar(df_mr2, "Qtd", motivo_r_col2, RED, 380, money=False),
                                    use_container_width=True, key="pc_20")
            panel_close()
        with cr2:
            panel_open("Top 10 clientes", tag="Qtd", icon="👥")
            if cliente_r_col and cliente_r_col in df_r.columns:
                df_clr = (df_r[df_r[cliente_r_col].str.strip() != ""]
                          .groupby(cliente_r_col).agg(Qtd=(cliente_r_col, "count"))
                          .reset_index().sort_values("Qtd", ascending=True).tail(10))
                if not df_clr.empty:
                    st.plotly_chart(make_hbar(df_clr, "Qtd", cliente_r_col, MIXED, 380, money=False),
                                    use_container_width=True, key="pc_21")
            panel_close()
        with cr3:
            panel_open("Top vendedores", tag="Qtd", icon="🧑‍💼")
            if nome_r_col and nome_r_col in df_r.columns:
                df_nomr = (df_r[df_r[nome_r_col].str.strip() != ""]
                           .groupby(nome_r_col).agg(Qtd=(nome_r_col, "count"))
                           .reset_index().sort_values("Qtd", ascending=True).tail(10))
                if not df_nomr.empty:
                    st.plotly_chart(make_hbar(df_nomr, "Qtd", nome_r_col, BLUE, 380, money=False),
                                    use_container_width=True, key="pc_22")
            panel_close()

        cr4, cr5 = st.columns([1, 2], gap="medium")
        with cr4:
            panel_open("Por praça", tag="Top 10", icon="🏙️")
            if praca_r_col and praca_r_col in df_r.columns:
                df_praca_r = (df_r[df_r[praca_r_col].str.strip() != ""]
                              .groupby(praca_r_col).agg(Qtd=(praca_r_col, "count"))
                              .reset_index().sort_values("Qtd", ascending=False).head(10))
                if not df_praca_r.empty:
                    fig_pr = make_donut(df_praca_r, praca_r_col, "Qtd", height=380,
                                        money=False, centro_rot="Reentregas")
                    st.plotly_chart(fig_pr, use_container_width=True, key="pc_23")
            panel_close()
        with cr5:
            panel_open("Ranking de motivos de transferência", tag="Consolidado", icon="📊")
            if motivo_r_col2 and motivo_r_col2 in df_r.columns:
                df_rkr = (df_r.groupby(motivo_r_col2).agg(Qtd=(motivo_r_col2, "count"))
                          .reset_index().sort_values("Qtd", ascending=False))
                tot_r = df_rkr["Qtd"].sum()
                df_rkr["%"] = (df_rkr["Qtd"] / tot_r * 100).round(1).astype(str) + "%" if tot_r > 0 else "0%"
                html_table(df_rkr.rename(columns={motivo_r_col2: "Motivo"}), max_rows=200, min_width=480)
            panel_close()


# ═════════════════════════════════════════════════════════════════════════════
# PÁGINA: DETALHES REENTREGAS
# ═════════════════════════════════════════════════════════════════════════════
elif pagina == "Detalhes Reentregas":
    if df_reent is None or len(df_reent) == 0:
        st.warning("Nenhum dado de reentregas disponível.")
    else:
        st.markdown('<div class="filters"><div class="filters-h"><span class="pip"></span>'
                    'Pesquisa — reentregas</div>', unsafe_allow_html=True)
        det_f1, det_f2, det_f3 = st.columns([3, 1.3, 1], gap="medium")
        usar_data_det = False
        dt_det_sel = None
        with det_f1:
            datas_det_ok = df_reent["_DATATRANSF_DT"].dropna()
            dt_col_det = reent_cols.get("DATATRANSF")
            col_label_det = dt_col_det if dt_col_det else "DTRANSF"
            if len(datas_det_ok) > 0:
                datas_det_unicas = sorted(datas_det_ok.dt.date.unique())
                opcoes_det = ["— Todas as datas —"] + [d.strftime("%d/%m/%Y") for d in datas_det_unicas]
                sel_det_str = st.selectbox(f"Data ({col_label_det})", opcoes_det, key="det_dtsel")
                if sel_det_str != "— Todas as datas —":
                    dt_det_sel = datetime.strptime(sel_det_str, "%d/%m/%Y").date()
                    usar_data_det = True
        with det_f2:
            st.markdown("<div style='height:26px'></div>", unsafe_allow_html=True)
            if st.button("Atualizar dados", use_container_width=True, type="primary", key="btn_det"):
                st.cache_data.clear()
                st.rerun()
        with det_f3:
            st.markdown("<div style='height:26px'></div>", unsafe_allow_html=True)
            if st.button("Limpar", use_container_width=True, key="btn_clear_det"):
                for k in ("det_dtsel", "det_cli", "det_nf", "det_ped", "det_placa"):
                    st.session_state.pop(k, None)
                st.rerun()

        ds1, ds2, ds3, ds4 = st.columns(4, gap="medium")
        with ds1:
            s_cli_r = st.text_input("Cliente", placeholder="Nome", key="det_cli")
        with ds2:
            s_nf_r = st.text_input("Nota (NUMNOTA)", placeholder="Número", key="det_nf")
        with ds3:
            s_ped_r = st.text_input("Pedido (NUMPED)", placeholder="Número", key="det_ped")
        with ds4:
            s_placa_r = st.text_input("Placa", placeholder="PLACAANT ou PLACAATUAL", key="det_placa")
        st.markdown('</div>', unsafe_allow_html=True)

        df_det_base = df_reent.copy()
        if usar_data_det and dt_det_sel:
            df_det_base = df_det_base[df_det_base["_DATATRANSF_DT"].dt.date == dt_det_sel]
            st.info(f"Data {dt_det_sel.strftime('%d/%m/%Y')} — {len(df_det_base)} registros nos gráficos e na tabela")
        else:
            st.info(f"Todos os períodos — {len(df_det_base)} registros. Use o filtro de data para detalhar.")

        REENT_DISPLAY = [
            ("DATATRANSF", "DTRANSF"), ("NUMPED", "NUMPED"), ("NUMNOTA", "NUMNOTA"),
            ("DTFAT", "DTFAT"), ("DTSAIDA", "DTSAIDA"), ("CLIENTE", "CLIENTE"), ("CODCLI", "CODCLI"),
            ("BAIRROENT", "BAIRROENT"), ("CODPRACA", "CODPRACA"), ("PRACA", "PRACA"), ("ROTA", "ROTA"),
            ("NUMCARANTERIOR", "CAR.ANT"), ("PLACAANT", "PLACAANT"),
            ("NOME_MOT_ANTERIOR", "MOT.ANT"), ("NOME_AJU_ANTERIOR", "AJU.ANT"),
            ("NUMCARATUAL", "CAR.ATUAL"), ("PLACAATUAL", "PLACAATUAL"),
            ("NOME_MOT_ATUAL", "MOT.ATUAL"), ("NOME_AJU_ATUAL", "AJU.ATUAL"),
            ("CODMOTIVO", "COD.MOTIVO"), ("MOTIVOTRANSF", "MOTIVO"),
            ("VLTOTGER", "VLTOTGER"), ("TOTPESO", "PESO"), ("NOME", "VENDEDOR"), ("CODUSU", "CODUSU"),
        ]
        cols_det_ok = [(reent_cols.get(k), label) for k, label in REENT_DISPLAY if reent_cols.get(k) is not None]
        df_det = df_det_base[[o for o, _ in cols_det_ok]].copy()
        df_det.columns = [label for _, label in cols_det_ok]

        # Nomes da aba NOMES aplicados às placas da reentrega
        if "PLACAANT" in df_det.columns:
            df_det["MOTORISTA (NOMES)"] = df_det["PLACAANT"].apply(
                lambda p: str(mapa_motorista.get(norm_placa(p), SEM_CADASTRO)).upper())
        if "PLACAATUAL" in df_det.columns:
            df_det["MOT.ATUAL (NOMES)"] = df_det["PLACAATUAL"].apply(
                lambda p: str(mapa_motorista.get(norm_placa(p), SEM_CADASTRO)).upper())

        if s_cli_r.strip() and "CLIENTE" in df_det.columns:
            df_det = df_det[df_det["CLIENTE"].str.contains(s_cli_r.strip(), case=False, na=False)]
        if s_nf_r.strip() and "NUMNOTA" in df_det.columns:
            df_det = df_det[df_det["NUMNOTA"].str.contains(s_nf_r.strip(), case=False, na=False)]
        if s_ped_r.strip() and "NUMPED" in df_det.columns:
            df_det = df_det[df_det["NUMPED"].str.contains(s_ped_r.strip(), case=False, na=False)]
        if s_placa_r.strip():
            mask_p = pd.Series([False] * len(df_det), index=df_det.index)
            for cp in ["PLACAANT", "PLACAATUAL"]:
                if cp in df_det.columns:
                    mask_p = mask_p | df_det[cp].str.contains(s_placa_r.strip(), case=False, na=False)
            df_det = df_det[mask_p]

        _det_placa_col = reent_cols.get("PLACAATUAL")
        _det_motivo_col = reent_cols.get("MOTIVOTRANSF")
        if not _det_placa_col:
            for _c in ["PLACA_ATUAL", "PLACAATUAL", "PLACA_ATU"]:
                if _c in df_det_base.columns:
                    _det_placa_col = _c
                    break
        if not _det_motivo_col:
            for _c in ["MOTIVOTRANSF", "MOTIVO_TRANSF"]:
                if _c in df_det_base.columns:
                    _det_motivo_col = _c
                    break

        gcol1, gcol2 = st.columns(2, gap="medium")
        with gcol1:
            panel_open("Reentregas por placa atual", tag="PLACAATUAL", icon="🚚")
            if _det_placa_col and _det_placa_col in df_det_base.columns:
                df_gplaca = (df_det_base[df_det_base[_det_placa_col].str.strip() != ""]
                             .groupby(_det_placa_col).agg(Qtd=(_det_placa_col, "count"))
                             .reset_index().sort_values("Qtd", ascending=False))
                if not df_gplaca.empty:
                    st.plotly_chart(make_bar_simple(df_gplaca, _det_placa_col, "Qtd", ramp(len(df_gplaca), 3, 6)),
                                    use_container_width=True, key="pc_24")
                else:
                    st.info("Sem placas para o filtro selecionado.")
            else:
                st.warning("Coluna PLACAATUAL não encontrada.")
            panel_close()
        with gcol2:
            panel_open("Reentregas por motorista atual", tag="Aba NOMES", icon="🧑‍✈️")
            if _det_placa_col and _det_placa_col in df_det_base.columns and mapa_motorista:
                _base_g = df_det_base[df_det_base[_det_placa_col].str.strip() != ""].copy()
                _base_g["_MOT"] = _base_g[_det_placa_col].apply(
                    lambda p: str(mapa_motorista.get(norm_placa(p), SEM_CADASTRO)).upper())
                df_gmotn = (_base_g.groupby("_MOT").agg(Qtd=("_MOT", "count"))
                            .reset_index().sort_values("Qtd", ascending=False))
                if not df_gmotn.empty:
                    st.plotly_chart(make_bar_simple(df_gmotn, "_MOT", "Qtd", ramp(len(df_gmotn), 3, 6, C_OK)),
                                    use_container_width=True, key="pc_25")
                else:
                    st.info("Sem motoristas para o filtro selecionado.")
            else:
                st.info("Cadastro de nomes indisponível.")
            panel_close()

        panel_open("Registros de reentrega", tag=f"{len(df_det)} linhas", icon="🔍")
        if len(df_det) == 0:
            st.warning("Nenhum registro encontrado. Ajuste a busca ou limpe os filtros.")
        else:
            html_table(df_det, min_width=1400)
            if len(df_det) > 500:
                st.caption(f"Exibindo as primeiras 500 de {len(df_det)} linhas.")
            st.markdown("<div style='margin-top:12px;'></div>", unsafe_allow_html=True)
            export_tabela(df_det, "reentregas", "dl_det", sheet_name="Reentregas")
        panel_close()

        with st.expander("Diagnóstico — colunas da planilha de reentregas"):
            cols_real = list(df_reent_raw.columns)
            st.write(f"**URL usada:** `{reent_url_usada}`")
            st.write(f"**Linhas:** {len(df_reent_raw)} · **Colunas:** {len(cols_real)}")
            st.write(f"**Colunas encontradas:** `{cols_real}`")
            st.write(f"PLACA_ATUAL → `{reent_cols.get('PLACAATUAL')}`")
            st.write(f"MOTIVOTRANSF → `{reent_cols.get('MOTIVOTRANSF')}`")
            st.write(f"DTRANSF → `{reent_cols.get('DATATRANSF')}`")
            st.write(f"PLACAANT → `{reent_cols.get('PLACAANT')}`")


# ═════════════════════════════════════════════════════════════════════════════
# PÁGINA: CAMPOS
# ═════════════════════════════════════════════════════════════════════════════
elif pagina == "Campos":
    st.markdown('<div class="filters"><div class="filters-h"><span class="pip"></span>'
                'Pesquisa — devoluções</div>', unsafe_allow_html=True)
    sr1, sr2, sr3, sr4, sr5 = st.columns(5, gap="medium")
    with sr1:
        s_cli = st.text_input("Cliente", placeholder="Nome", key="sc_cli")
    with sr2:
        s_nf = st.text_input("Nota de venda", placeholder="Número", key="sc_nf")
    with sr3:
        s_ped = st.text_input("Código do cliente", placeholder="CODCLI", key="sc_ped")
    with sr4:
        s_placa2 = st.text_input("Placa", placeholder="Ex.: NPB1J08", key="sc_placa")
    with sr5:
        s_mot_nome = st.text_input("Motorista / Entregador", placeholder="Nome", key="sc_motnome")
    st.markdown('</div>', unsafe_allow_html=True)

    CAMPOS = [
        (COL_DTENTREGA, "DTENT"), (COL_DTSAIDA, "DTSAIDA"), (COL_NF_VENDA, "NOTA_VENDA"),
        (COL_NOTA_DEV, "NOTA_DEVOLUCAO"), (COL_NUMCAR, "NUMCAR"), (COL_PLACA, "PLACA"),
        (COL_MOT_NOME, "MOTORISTA_NOMES"), (COL_ENT_NOME, "ENTREGADOR_NOMES"),
        (COL_DESTINO, "DESTINO"), (COL_MOTIVO, "MOTIVO"), (COL_CODCLI, "CODCLI"),
        (COL_CLIENTE, "CLIENTE"), (COL_MOTORISTA, "MOTORISTA"), (COL_VENDEDOR, "NOMERCA"),
        (COL_DEVOLUCION, "NOMEFUNC"), (COL_SUPERVISOR, "SUPERVISOR"), (COL_TIPO_MERC, "TIPO_MERCADO"),
    ]
    cols_ok = [(o, a) for o, a in CAMPOS if o is not None]
    df_campos = df[[o for o, _ in cols_ok]].copy()
    df_campos.columns = [a for _, a in cols_ok]

    if s_cli.strip() and "CLIENTE" in df_campos.columns:
        df_campos = df_campos[df_campos["CLIENTE"].str.contains(s_cli.strip(), case=False, na=False)]
    if s_nf.strip() and "NOTA_VENDA" in df_campos.columns:
        df_campos = df_campos[df_campos["NOTA_VENDA"].str.contains(s_nf.strip(), case=False, na=False)]
    if s_ped.strip() and "CODCLI" in df_campos.columns:
        df_campos = df_campos[df_campos["CODCLI"].str.contains(s_ped.strip(), case=False, na=False)]
    if s_placa2.strip() and "PLACA" in df_campos.columns:
        df_campos = df_campos[df_campos["PLACA"].str.contains(s_placa2.strip(), case=False, na=False)]
    if s_mot_nome.strip():
        _t = s_mot_nome.strip()
        _m = pd.Series([False] * len(df_campos), index=df_campos.index)
        for _c in ["MOTORISTA_NOMES", "ENTREGADOR_NOMES"]:
            if _c in df_campos.columns:
                _m = _m | df_campos[_c].str.contains(_t, case=False, na=False)
        df_campos = df_campos[_m]

    panel_open("Campos das devoluções", tag=f"{len(df_campos)} registros", icon="🗂️")
    if len(df_campos) == 0:
        st.warning("Nenhum registro encontrado. Ajuste a busca ou limpe os filtros.")
    else:
        html_table(df_campos, min_width=1300)
        if len(df_campos) > 500:
            st.caption(f"Exibindo as primeiras 500 de {len(df_campos)} linhas.")
        st.markdown("<div style='margin-top:12px;'></div>", unsafe_allow_html=True)
        export_tabela(df_campos, "campos_devolucoes", "dl_campos", sheet_name="Campos")
    panel_close()


# ═════════════════════════════════════════════════════════════════════════════
# PÁGINA: DADOS COMPLETOS
# ═════════════════════════════════════════════════════════════════════════════
elif pagina == "Dados Completos":
    display_cols = [c for c in actual_cols if not c.startswith("_")]
    # As duas colunas derivadas da aba NOMES entram junto com as originais
    display_cols = display_cols + [c for c in [COL_MOT_NOME, COL_ENT_NOME] if c not in display_cols]

    st.markdown('<div class="filters"><div class="filters-h"><span class="pip"></span>'
                'Exibição da tabela</div>', unsafe_allow_html=True)
    d1, d2, d3 = st.columns(3, gap="medium")
    sort_opts = [VALOR_COL] + [c for c in [COL_DTSAIDA, COL_DTENTREGA, COL_CLIENTE, COL_MOTIVO,
                                           COL_PLACA, COL_MOT_NOME, COL_ENT_NOME] if c]
    with d1:
        sort_col = st.selectbox("Ordenar por", sort_opts)
    with d2:
        sort_asc = st.radio("Direção", ["Crescente", "Decrescente"], horizontal=True) == "Crescente"
    with d3:
        n_rows = st.selectbox("Máximo de linhas", [50, 100, 250, 500, 1000, "Todos"])
    st.markdown('</div>', unsafe_allow_html=True)

    df_sorted = df.sort_values(sort_col, ascending=sort_asc)
    if n_rows != "Todos":
        df_sorted = df_sorted.head(int(n_rows))
    disp = df_sorted[display_cols]

    panel_open("Base completa de devoluções", tag=f"{len(df)} registros no filtro", icon="📑")
    html_table(disp, min_width=1200)
    st.caption(f"Exibindo {min(len(disp), 500)} de {len(df)} registros.")
    st.markdown("<div style='margin-top:12px;'></div>", unsafe_allow_html=True)
    e1, e2 = st.columns([2, 1])
    with e1:
        export_tabela(df[display_cols], "devolucoes", "dl_completos", sheet_name="Devoluções")
    with e2:
        if st.button("Atualizar dados", use_container_width=True, type="primary", key="btn_upd_dados"):
            st.cache_data.clear()
            st.rerun()
    panel_close()


# ═════════════════════════════════════════════════════════════════════════════
# PÁGINA: CLIENTES
# ═════════════════════════════════════════════════════════════════════════════
elif pagina == "Clientes":
    if not COL_CLIENTE:
        st.warning("Coluna CLIENTE não encontrada na planilha.")
    else:
        df_cli_all = (df[df[COL_CLIENTE].str.strip() != ""]
                      .groupby(COL_CLIENTE).agg(Valor=(VALOR_COL, "sum"), Qtd=(VALOR_COL, "count"))
                      .reset_index().sort_values("Valor", ascending=False))
        st.markdown(
            '<div class="kpi-grid">'
            + kpi("👥", "Clientes únicos", f"{total_clientes}", "Com devolução no filtro", "#34d399")
            + kpi("💰", "Valor total", fmt_brl(total_val), "Soma de VLTOTAL", "#22d3ee")
            + kpi("📄", "Devoluções", f"{total_notas}", "Notas no filtro", "#a78bfa")
            + kpi("📈", "Ticket médio", fmt_brl(ticket_medio), "Valor por devolução", "#fbbf24")
            + kpi("🏆", "Maior cliente",
                  fmt_brl0(df_cli_all["Valor"].iloc[0]) if not df_cli_all.empty else "R$ 0",
                  str(df_cli_all[COL_CLIENTE].iloc[0])[:26] if not df_cli_all.empty else "—", "#fb923c")
            + '</div>', unsafe_allow_html=True)

        cA, cB = st.columns([1.4, 1], gap="medium")
        with cA:
            panel_open("Clientes com maior valor devolvido", tag="Top 15", icon="👥")
            df_top15 = df_cli_all.sort_values("Valor", ascending=True).tail(15)
            if not df_top15.empty:
                st.plotly_chart(make_hbar(df_top15, "Valor", COL_CLIENTE, MIXED, 560), use_container_width=True, key="pc_26")
            else:
                st.info("Sem clientes no filtro atual.")
            panel_close()
        with cB:
            panel_open("Clientes por quantidade de notas", tag="Top 15", icon="📄")
            df_topq = df_cli_all.sort_values("Qtd", ascending=True).tail(15)
            if not df_topq.empty:
                st.plotly_chart(make_hbar(df_topq, "Qtd", COL_CLIENTE, BLUE, 560, money=False),
                                use_container_width=True, key="pc_27")
            panel_close()

        panel_open("Detalhamento por cliente", tag=f"{len(df_cli_all)} clientes", icon="📋")
        tb = df_cli_all.copy()
        tb["Valor total"] = tb["Valor"].apply(fmt_brl)
        tb["% do total"] = (tb["Valor"] / total_val * 100).round(1).astype(str) + "%" if total_val > 0 else "0%"
        tb["Ticket médio"] = (tb["Valor"] / tb["Qtd"]).apply(fmt_brl)
        tb = tb.rename(columns={COL_CLIENTE: "Cliente", "Qtd": "Notas"})
        _tb_cli = tb[["Cliente", "Notas", "Valor total", "Ticket médio", "% do total"]]
        html_table(_tb_cli, min_width=760)
        st.markdown("<div style='margin-top:12px;'></div>", unsafe_allow_html=True)
        export_tabela(_tb_cli, "clientes", "dl_cli", sheet_name="Clientes")
        panel_close()


# ═════════════════════════════════════════════════════════════════════════════
# PÁGINA: MOTIVOS
# ═════════════════════════════════════════════════════════════════════════════
elif pagina == "Motivos":
    if not COL_MOTIVO:
        st.warning("Coluna MOTIVO não encontrada na planilha.")
    else:
        df_mot_all = (df[df[COL_MOTIVO].str.strip() != ""]
                      .groupby(COL_MOTIVO).agg(Valor=(VALOR_COL, "sum"), Qtd=(VALOR_COL, "count"))
                      .reset_index().sort_values("Valor", ascending=False))
        st.markdown(
            '<div class="kpi-grid">'
            + kpi("❗", "Motivos distintos", f"{len(df_mot_all)}", "No filtro atual", "#f87171")
            + kpi("💰", "Valor total", fmt_brl(total_val), "Soma de VLTOTAL", "#22d3ee")
            + kpi("📄", "Devoluções", f"{total_notas}", "Notas no filtro", "#a78bfa")
            + kpi("🥇", "Motivo líder",
                  fmt_brl0(df_mot_all["Valor"].iloc[0]) if not df_mot_all.empty else "R$ 0",
                  str(df_mot_all[COL_MOTIVO].iloc[0])[:26] if not df_mot_all.empty else "—", "#fbbf24")
            + kpi("📈", "Ticket médio", fmt_brl(ticket_medio), "Valor por devolução", "#34d399")
            + '</div>', unsafe_allow_html=True)

        panel_open("Motivos — valor e quantidade", tag="Consolidado", icon="❗")
        if not df_mot_all.empty:
            fig_m = make_combo_chart(df_mot_all, COL_MOTIVO, "Valor", "Qtd", "", "", ramp(len(df_mot_all), 3, 6))
            fig_m.update_layout(height=max(440, min(len(df_mot_all) * 58, 680)),
                                margin=dict(t=54, b=120, l=10, r=30))
            fig_m.update_xaxes(tickangle=-35, automargin=True)
            st.plotly_chart(fig_m, use_container_width=True, key="pc_28")
        else:
            st.info("Sem motivos no filtro atual.")
        panel_close()

        m1, m2 = st.columns([1, 1.4], gap="medium")
        with m1:
            panel_open("Participação por motivo", tag="Top 10", icon="🥧")
            df_pie = df_mot_all.head(10)
            if not df_pie.empty:
                fig_pm = make_donut(df_pie, COL_MOTIVO, "Valor", height=430,
                                    money=True, centro_rot="Devolvido")
                st.plotly_chart(fig_pm, use_container_width=True, key="pc_29")
            panel_close()
        with m2:
            panel_open("Ranking completo", tag=f"{len(df_mot_all)} motivos", icon="📊")
            tbm = df_mot_all.copy()
            tbm["Valor total"] = tbm["Valor"].apply(fmt_brl)
            tbm["% do total"] = (tbm["Valor"] / total_val * 100).round(1).astype(str) + "%" if total_val > 0 else "0%"
            tbm = tbm.rename(columns={COL_MOTIVO: "Motivo", "Qtd": "Ocorrências"})
            _tb_mot = tbm[["Motivo", "Ocorrências", "Valor total", "% do total"]]
            html_table(_tb_mot, min_width=620)
            st.markdown("<div style='margin-top:12px;'></div>", unsafe_allow_html=True)
            export_tabela(_tb_mot, "motivos", "dl_motivos", sheet_name="Motivos")
            panel_close()


# ═════════════════════════════════════════════════════════════════════════════
# PÁGINA: VEÍCULOS
# ═════════════════════════════════════════════════════════════════════════════
elif pagina == "Veículos":
    if not COL_PLACA:
        st.warning("Coluna PLACA não encontrada na planilha.")
    else:
        _agg_v = dict(Valor=(VALOR_COL, "sum"), Qtd=(VALOR_COL, "count"),
                      Motorista=(COL_MOT_NOME, "first"), Entregador=(COL_ENT_NOME, "first"))
        df_pl_all = (df[df[COL_PLACA].str.strip() != ""]
                     .groupby(COL_PLACA).agg(**_agg_v)
                     .reset_index().sort_values("Valor", ascending=False))
        st.markdown(
            '<div class="kpi-grid">'
            + kpi("🚚", "Veículos únicos", f"{total_placas}", "Placas no filtro", "#fb923c")
            + kpi("💰", "Valor total", fmt_brl(total_val), "Soma de VLTOTAL", "#22d3ee")
            + kpi("📄", "Devoluções", f"{total_notas}", "Notas no filtro", "#a78bfa")
            + kpi("📊", "Média por veículo",
                  fmt_brl(total_val / total_placas) if total_placas > 0 else "R$ 0,00",
                  "Valor médio por placa", "#34d399")
            + kpi("🥇", "Maior valor",
                  fmt_brl0(df_pl_all["Valor"].iloc[0]) if not df_pl_all.empty else "R$ 0",
                  (f"{df_pl_all[COL_PLACA].iloc[0]} · {str(df_pl_all['Motorista'].iloc[0])[:18]}"
                   if not df_pl_all.empty else "—"), "#f87171")
            + '</div>', unsafe_allow_html=True)

        panel_open("Devoluções por veículo — valor e quantidade", tag="Consolidado", icon="🚚")
        if not df_pl_all.empty:
            st.plotly_chart(
                make_combo_chart(df_pl_all, COL_PLACA, "Valor", "Qtd", "", "", ramp(len(df_pl_all)),
                                 customdata=df_pl_all[["Motorista", "Entregador"]].values,
                                 extra_hover="<br>🧑‍✈️ %{customdata[0]}<br>📦 %{customdata[1]}"),
                use_container_width=True, key="pc_30")
        else:
            st.info("Sem veículos no filtro atual.")
        panel_close()

        v1, v2 = st.columns([1, 1.4], gap="medium")
        with v1:
            panel_open("Veículos por quantidade de notas", tag="Top 12", icon="📄")
            df_pq = df_pl_all.sort_values("Qtd", ascending=True).tail(12)
            if not df_pq.empty:
                st.plotly_chart(make_hbar(df_pq, "Qtd", COL_PLACA, BLUE, 460, money=False),
                                use_container_width=True, key="pc_31")
            panel_close()
        with v2:
            panel_open("Detalhamento por veículo", tag=f"{len(df_pl_all)} placas", icon="📋")
            tbv = df_pl_all.copy()
            tbv["Valor total"] = tbv["Valor"].apply(fmt_brl)
            tbv["Ticket médio"] = (tbv["Valor"] / tbv["Qtd"]).apply(fmt_brl)
            tbv["% do total"] = (tbv["Valor"] / total_val * 100).round(1).astype(str) + "%" if total_val > 0 else "0%"
            tbv = tbv.rename(columns={COL_PLACA: "Placa", "Qtd": "Notas"})
            _tb_vei = tbv[["Placa", "Motorista", "Entregador", "Notas", "Valor total",
                           "Ticket médio", "% do total"]]
            html_table(_tb_vei, min_width=980)
            st.markdown("<div style='margin-top:12px;'></div>", unsafe_allow_html=True)
            export_tabela(_tb_vei, "veiculos", "dl_veic", sheet_name="Veículos")
            panel_close()


# ═════════════════════════════════════════════════════════════════════════════
# PÁGINA: RELATÓRIOS
# ═════════════════════════════════════════════════════════════════════════════
elif pagina == "Relatórios":
    st.markdown(
        '<div class="kpi-grid">'
        + kpi("📄", "Devoluções no filtro", f"{total_notas}", "Prontas para exportar", "#22d3ee")
        + kpi("💰", "Valor total", fmt_brl(total_val), "Soma de VLTOTAL", "#34d399")
        + kpi("🔄", "Reentregas", f"{len(df_reent) if df_reent is not None else 0}", "Registros carregados", "#fbbf24")
        + kpi("🧑‍✈️", "Motoristas", f"{total_motoristas}", "No filtro atual", "#a78bfa")
        + kpi("🚚", "Veículos únicos", f"{total_placas}", "No filtro atual", "#fb923c")
        + '</div>', unsafe_allow_html=True)

    panel_open("Exportações", tag="Excel formatado como tabela", icon="📤")
    st.markdown('<p style="font-size:0.82rem;color:#7c8ea8;margin-bottom:14px;">'
                'Os arquivos respeitam os filtros aplicados no topo da página. O Excel já sai como '
                'tabela: cabeçalho fixo, filtro em cada coluna, valores em formato de moeda e '
                'larguras ajustadas.</p>', unsafe_allow_html=True)
    display_cols_rel = [c for c in actual_cols if not c.startswith("_")]
    display_cols_rel = display_cols_rel + [c for c in [COL_MOT_NOME, COL_ENT_NOME]
                                           if c not in display_cols_rel]

    r1, r2 = st.columns(2, gap="large")
    with r1:
        st.markdown('<p style="font-size:0.78rem;color:#c3d0e0;font-weight:600;'
                    'margin:0 0 6px;">Devoluções filtradas</p>', unsafe_allow_html=True)
        export_tabela(df[display_cols_rel], "devolucoes", "rel_dev", sheet_name="Devoluções")

        st.markdown('<p style="font-size:0.78rem;color:#c3d0e0;font-weight:600;'
                    'margin:16px 0 6px;">Resumo por motivo</p>', unsafe_allow_html=True)
        if COL_MOTIVO:
            _rm = (df.groupby(COL_MOTIVO).agg(Qtd=(VALOR_COL, "count"), Valor=(VALOR_COL, "sum"))
                   .reset_index().sort_values("Valor", ascending=False))
            _rm_x = _rm.rename(columns={COL_MOTIVO: "Motivo", "Qtd": "Ocorrências"})
            _rm_x["Valor"] = _rm_x["Valor"].apply(fmt_brl)
            export_tabela(_rm_x, "resumo_motivos", "rel_mot", sheet_name="Motivos")
        else:
            st.caption("Coluna de motivo indisponível.")

    with r2:
        st.markdown('<p style="font-size:0.78rem;color:#c3d0e0;font-weight:600;'
                    'margin:0 0 6px;">Resumo por motorista</p>', unsafe_allow_html=True)
        _re = agrupa_equipe(df, COL_MOT_NOME)
        if not _re.empty:
            _re_x = _re.rename(columns={COL_MOT_NOME: "Motorista", "Qtd": "Notas"})
            _re_x["Valor"] = _re_x["Valor"].apply(fmt_brl)
            export_tabela(_re_x, "resumo_motoristas", "rel_mtr", sheet_name="Motoristas")
        else:
            st.caption("Sem motoristas vinculados.")

        st.markdown('<p style="font-size:0.78rem;color:#c3d0e0;font-weight:600;'
                    'margin:16px 0 6px;">Reentregas completas</p>', unsafe_allow_html=True)
        if df_reent is not None and len(df_reent) > 0:
            _cols_r = [c for c in df_reent.columns if not c.startswith("_")]
            export_tabela(df_reent[_cols_r], "reentregas", "rel_reent", sheet_name="Reentregas")
        else:
            st.caption("Reentregas indisponíveis no momento.")
    panel_close()

    panel_open("Resumo consolidado", tag="Visão do filtro atual", icon="📋")
    linhas = []
    if COL_MOTIVO:
        _t = (df.groupby(COL_MOTIVO).agg(Qtd=(VALOR_COL, "count"), Valor=(VALOR_COL, "sum"))
              .reset_index().sort_values("Valor", ascending=False).head(10))
        for _, r in _t.iterrows():
            linhas.append({"Dimensão": "Motivo", "Item": r[COL_MOTIVO], "Qtd": r["Qtd"],
                           "Valor": fmt_brl(r["Valor"])})
    if COL_PLACA:
        _t = (df[df[COL_PLACA].str.strip() != ""].groupby(COL_PLACA)
              .agg(Qtd=(VALOR_COL, "count"), Valor=(VALOR_COL, "sum"))
              .reset_index().sort_values("Valor", ascending=False).head(10))
        for _, r in _t.iterrows():
            linhas.append({"Dimensão": "Veículo", "Item": r[COL_PLACA], "Qtd": r["Qtd"],
                           "Valor": fmt_brl(r["Valor"])})
    _tm_res = agrupa_equipe(df, COL_MOT_NOME).head(10)
    for _, r in _tm_res.iterrows():
        linhas.append({"Dimensão": "Motorista", "Item": r[COL_MOT_NOME], "Qtd": r["Qtd"],
                       "Valor": fmt_brl(r["Valor"])})
    if linhas:
        _df_res = pd.DataFrame(linhas)
        html_table(_df_res, min_width=560)
        st.markdown("<div style='margin-top:12px;'></div>", unsafe_allow_html=True)
        export_tabela(_df_res, "resumo_consolidado", "rel_res", sheet_name="Resumo")
    else:
        st.info("Nada a resumir com os filtros atuais.")
    panel_close()


# ═════════════════════════════════════════════════════════════════════════════
# PÁGINA: CONFIGURAÇÕES
# ═════════════════════════════════════════════════════════════════════════════
elif pagina == "Configurações":
    panel_open("Fonte de dados", tag="Google Sheets", icon="🔌")
    st.markdown(f'<p style="font-size:0.82rem;color:#b9c8dc;">Planilha: <code>{SHEET_ID}</code><br>'
                f'Cache: 60 segundos · Última sincronização: {_sync}</p>', unsafe_allow_html=True)
    cfg1, cfg2 = st.columns([1, 3])
    with cfg1:
        if st.button("Atualizar dados", use_container_width=True, type="primary", key="btn_cfg"):
            st.cache_data.clear()
            st.rerun()
    panel_close()

    panel_open("Colunas detectadas — devoluções", tag=f"{len(actual_cols)} colunas", icon="🧩")
    st.write(f"**Colunas:** `{actual_cols}`")
    st.write(f"Valor = `{VALOR_COL}` · Placa = `{COL_PLACA}` · Motivo = `{COL_MOTIVO}`")
    st.write(f"Cliente = `{COL_CLIENTE}` · Devolucionista = `{COL_DEVOLUCION}`")
    st.write(f"Data de entrada (DTENT) = `{COL_DTENTREGA}` · Data de saída = `{COL_DTSAIDA}`")
    st.write(f"Supervisor = `{COL_SUPERVISOR}` · Destino = `{COL_DESTINO}` · Nota de devolução = `{COL_NOTA_DEV}`")
    st.write(f"Pedido (NUMPED) = `{COL_NUMPED}`")
    st.write(f"Registros com valor maior que zero: {(df_raw[VALOR_COL] > 0).sum()}")
    panel_close()

    panel_open("Cadastro de equipe — aba NOMES", tag="Diagnóstico", icon="🧑‍✈️")
    if df_nomes_norm is not None:
        st.write(f"**URL usada:** `{nomes_url_usada}`")
        st.write(f"**Linhas no cadastro:** {len(df_nomes_norm)} · "
                 f"**Colunas:** `{[c for c in df_nomes_norm.columns if c != '_PLACA_KEY']}`")
        st.write(f"Placas distintas nas devoluções: {len(placas_base)} · "
                 f"Placas cruzadas com sucesso: {len(placas_base & placas_cadastradas)}")
        if placas_sem_cadastro:
            st.warning(f"Placas sem cadastro na aba NOMES ({len(placas_sem_cadastro)}): "
                       f"{', '.join(placas_sem_cadastro[:40])}"
                       f"{'…' if len(placas_sem_cadastro) > 40 else ''}")
        else:
            st.success("Todas as placas das devoluções têm motorista/entregador cadastrado.")
        with st.expander("Ver cadastro completo"):
            html_table(df_nomes_norm.drop(columns=["_PLACA_KEY"]), min_width=700)
    else:
        st.warning("A aba NOMES não foi carregada nesta sessão.")
        if nomes_load_error:
            st.caption(nomes_load_error)
        with st.expander("URLs tentadas"):
            for u in NOMES_URLS:
                st.code(u)
    panel_close()

    panel_open("Colunas detectadas — reentregas", tag="Diagnóstico", icon="🧩")
    if df_reent is not None:
        st.write(f"**URL usada:** `{reent_url_usada}`")
        st.write(f"**Colunas:** `{list(df_reent_raw.columns)}`")
        st.write(f"Valor = `{REENT_VALOR_COL}` · PLACAANT = `{reent_cols.get('PLACAANT')}` · "
                 f"PLACAATUAL = `{reent_cols.get('PLACAATUAL')}` · MOTIVOTRANSF = `{reent_cols.get('MOTIVOTRANSF')}`")
    else:
        st.warning("Reentregas não carregadas nesta sessão.")
        if reent_load_error:
            st.caption(reent_load_error)
    panel_close()

    panel_open("Sobre esta versão", tag="Interface v3.0", icon="ℹ️")
    st.markdown(
        '<p style="font-size:0.82rem;color:#7c8ea8;line-height:1.8;">'
        'Redesign visual: menu lateral fixo, cabeçalho com status de sincronização, painel de filtros compacto, '
        'cards de indicadores redesenhados e gráficos com eixos discretos.<br>'
        'v2.1: painel "Pedidos com devolução" no Dashboard, com busca, ordenação e exportação em CSV.<br>'
        'v2.2: leitura da aba <b>NOMES</b> (PLACA · MOTORISTA · ENTREGADOR), cruzamento com a placa das '
        'devoluções, gráficos por motorista e por entregador, filtro global de motorista, nova página '
        '<b>Equipe</b> e nomes no tooltip dos gráficos por placa.<br>'
        'v3.0: reformulação visual completa dos gráficos — tokens de design únicos, linguagem de cor por significado (vermelho crítico, laranja intermediário, amarelo quantidade, azul informativo, verde positivo, cinza neutro), números maiores em etiquetas escuras que nunca se sobrepõem, eixos laterais removidos, roscas no lugar das pizzas e dimensionamento responsivo por quantidade de categorias.<br>''Nenhuma regra de cálculo, consulta, filtro ou fonte de dados anterior foi alterada.</p>',
        unsafe_allow_html=True)
    panel_close()
